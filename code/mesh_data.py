import json
import os
from PIL import Image, ImageDraw, ImageFilter, ImageFont
import random
import math
import numpy as np
import copy

import fitz
import pickle
import json
from numpy.core.fromnumeric import repeat, reshape
from numpy.lib.function_base import diff
import yaml
from multiprocessing import Pool, Manager, managers
import cv2
from openpyxl import Workbook

Image.MAX_IMAGE_PIXELS = 2300000000


def mesh_data_test(test_img_path, crop_width, crop_height, save_small_img_parent_path=None, moving_radio=0.5):
    # 将输入的大图片切割为多个小图片，用于模型测试
    big_img = Image.open(test_img_path).convert('RGB')
    total_width, total_height = big_img.size[0], big_img.size[1]
    moving_step_width, moving_step_height = int(crop_width*moving_radio), int(crop_height*moving_radio)  # 近似的滑动步长
    num_width, num_height = math.ceil((total_width-crop_width)/moving_step_width) + 1, math.ceil(
        (total_height-crop_height)/moving_step_height) + 1  # 宽度和高度上分别划分为多少个子图
    moving_step_width, moving_step_height = math.ceil((total_width-crop_width)/(
        num_width-1)), math.ceil((total_height-crop_height)/(num_height-1))  # 实际滑动步长
    top_left_points = [[x*moving_step_width, y*moving_step_height]
                       for y in range(num_height) for x in range(num_width)]

    save_small_imgs = []
    mesh_info = {}
    mesh_info['total_width'] = total_width
    mesh_info['total_height'] = total_height
    mesh_info['num_width'] = num_width
    mesh_info['num_height'] = num_height
    mesh_info['moving_step_width'] = moving_step_width
    mesh_info['moving_step_height'] = moving_step_height
    mesh_info['crop_width'] = crop_width
    mesh_info['crop_height'] = crop_height
    reality_top_left_points = []
    for index, (top_left_x, top_left_y) in enumerate(top_left_points):
        bottom_right_x = min(total_width, top_left_x+crop_width)
        bottom_right_y = min(total_height, top_left_y+crop_height)
        top_left_x = bottom_right_x - crop_width
        top_left_y = bottom_right_y - crop_height
        small_img = big_img.crop(
            (top_left_x, top_left_y, bottom_right_x, bottom_right_y))
        save_small_imgs.append(small_img)
        file_name, suffix = os.path.split(
            test_img_path)[-1].split('.')[0], os.path.split(test_img_path)[-1].split('.')[1]
        if save_small_img_parent_path:
            if not os.path.exists(save_small_img_parent_path):
                os.makedirs(save_small_img_parent_path)
            # small_img.save(os.path.join(save_small_img_parent_path,
                                        # '{}_{}.{}'.format(file_name, index, suffix)))
        reality_top_left_points.append([top_left_x, top_left_y])

    mesh_info['top_left_points'] = reality_top_left_points
    # mesh_info['data'] = save_small_imgs
    if save_small_img_parent_path:
        with open(os.path.join(save_small_img_parent_path, 'top_left_point.pkl'), 'wb') as f:
            pickle.dump(mesh_info, f)

    return big_img, save_small_imgs, mesh_info


def py_cpu_nms(dets, thresh):
    x1 = dets[:, 1] - dets[:, 3]/2
    y1 = dets[:, 2] - dets[:, 4]/2
    x2 = dets[:, 1] + dets[:, 3]/2
    y2 = dets[:, 2] + dets[:, 4]/2
    
    areas = (y2-y1+1) * (x2-x1+1)
    scores = dets[:, 5]
    keep = []
    
    index = scores.argsort()[::-1]
    
    while index.size >0:
        i = index[0]       # every time the first is the biggst, and add it directly
        keep.append(i)
        x11 = np.maximum(x1[i], x1[index[1:]])    # calculate the points of overlap 
        y11 = np.maximum(y1[i], y1[index[1:]])
        x22 = np.minimum(x2[i], x2[index[1:]])
        y22 = np.minimum(y2[i], y2[index[1:]])
        w = np.maximum(0, x22-x11+1)    # the weights of overlap
        h = np.maximum(0, y22-y11+1)    # the height of overlap
        overlaps = w*h
        tem_data = np.ones_like(areas[index[1:]]) * areas[i]
        c = np.concatenate((np.reshape(tem_data, (-1, 1)), np.reshape(areas[index[1:]], (-1, 1))), axis=1)
        a = np.min(c, axis=1)
        ious = overlaps / a
        # ious = overlaps / np.min(areas[i], areas[index[1:]])
        idx = np.where(ious<=thresh)[0]
        index = index[idx+1]   # because index start from 1
        
    return keep


def merge_one_subplot(rectangle_1, rectangle_2, big_img_object_region, t):
    label_1, x_1_center, y_1_center, width_1, height_1, probability_1 = rectangle_1
    label_2, x_2_center, y_2_center, width_2, height_2, probability_2 = rectangle_2
    if abs(x_1_center-x_2_center)>(width_1/2+width_2/2) or abs(y_1_center-y_2_center)>(height_1/2+height_2/2):
        return big_img_object_region, 0
    x_1_0, y_1_0, x_1_1, y_1_1 = x_1_center-width_1/2, y_1_center - \
        height_1/2, x_1_center+width_1/2, y_1_center+height_1/2
    x_2_0, y_2_0, x_2_1, y_2_1 = x_2_center-width_2/2, y_2_center - \
        height_2/2, x_2_center+width_2/2, y_2_center+height_2/2
    # 求交集矩形
    x_0 = max(x_1_0, x_2_0)
    y_0 = max(y_1_0, y_2_0)
    x_1 = min(x_1_1, x_2_1)
    y_1 = min(y_1_1, y_2_1)

    merge_once = 0
    if (x_1-x_0) > 0 and (y_1-y_0) > 0:
        s_share = (x_1-x_0) * (y_1-y_0)
        iou_1 = s_share / (x_1_1-x_1_0)*(y_1_1-y_1_0)
        iou_2 = s_share / (x_2_1-x_2_0)*(y_2_1-y_2_0)
        if max(iou_1, iou_2) > t:
            # # 求并集矩形
            # merge_x_0 = min(x_1_0, x_2_0)
            # merge_y_0 = min(y_1_0, y_2_0)
            # merge_x_1 = max(x_1_1, x_2_1)
            # merge_y_1 = max(y_1_1, y_2_1)
            # merge_object_region = [
            #     label_1, (merge_x_0+merge_x_1)/2, (merge_y_0+merge_y_1)/2, merge_x_1-merge_x_0, merge_y_1-merge_y_0, (probability_1+probability_2)/2]
            # big_img_object_region.append(merge_object_region)
            if rectangle_1 in big_img_object_region and probability_1<=probability_2:
                big_img_object_region.remove(rectangle_1)
            if rectangle_2 in big_img_object_region and probability_2<probability_1:
                big_img_object_region.remove(rectangle_2)
            merge_once = 1
    return big_img_object_region, merge_once


def merge_objects_region(top_left_points, objects_region, t=0.5):
    # 将每个子图中检测到的目标合并到大图中
    # top_left_points, objects_region需与原图中尺度一致
    label_info = label_get_info(r'data/config.json')
    assert len(top_left_points) == len(objects_region)
    big_img_object_region = []
    for top_left_point, object_region in zip(top_left_points, objects_region):
        top_left_x, top_left_y = top_left_point
        if len(object_region) > 0:
            for label, x_0, y_0, x_1, y_1, probability in object_region:  # 单个子图内的多个目标
                center_x = (x_0 + x_1)/2 + top_left_x
                center_y = (y_0 + y_1)/2 + top_left_y
                width = x_1 - x_0
                height = y_1 - y_0

                if [label, center_x, center_y, width, height, probability, type] not in big_img_object_region:
                    big_img_object_region.append([label, center_x, center_y, width, height, probability])
    
    for item in copy.deepcopy(big_img_object_region):
        label, center_x, center_y, width, height, probability = item
        if probability < 0.5:
            big_img_object_region.remove(item)
    
    # while True: # 去掉相同标签的重叠区域
    #     merge_times_sum = 0
    #     rectangle_1_list = copy.deepcopy(big_img_object_region)
    #     for index, rectangle_1 in enumerate(rectangle_1_list):
    #         for rectangle_2 in rectangle_1_list[index+1:]:
    #             if rectangle_1[0] != rectangle_2[0]:
    #                 continue
    #             big_img_object_region, merge_once = merge_one_subplot(rectangle_1, rectangle_2, big_img_object_region, t)
    #             merge_times_sum = merge_times_sum + merge_once
    #             if merge_once==1:   # 每个区域在每轮合并中仅被合并一次
    #                 break
    #     if merge_times_sum==0:
    #         break

    big_img_object_region_cell = []
    big_img_object_region_conn = []
    big_img_object_region_sign = []
    for item in big_img_object_region:
        label, center_x, center_y, width, height, probability = item
        if label_info[label]['type']=='cell':
            big_img_object_region_cell.append(item)
        elif label_info[label]['type']=='connection':
            big_img_object_region_conn.append(item)
        elif label_info[label]['type']=='sign':
            big_img_object_region_sign.append(item)

    if len(big_img_object_region_cell)>0:
        select_index = py_cpu_nms(np.array(big_img_object_region_cell), t)
        big_img_object_region_cell = np.array(big_img_object_region_cell)[select_index]
    else:
        big_img_object_region_cell = np.array(big_img_object_region_cell)
    big_img_object_region_cell = np.reshape(big_img_object_region_cell, (-1, 6))

    if len(big_img_object_region_conn):
        select_index = py_cpu_nms(np.array(big_img_object_region_conn), t)
        big_img_object_region_conn = np.array(big_img_object_region_conn)[select_index]
    else:
        big_img_object_region_conn = np.array(big_img_object_region_conn)
    big_img_object_region_conn = np.reshape(big_img_object_region_conn, (-1, 6))

    if len(big_img_object_region_sign):
        select_index = py_cpu_nms(np.array(big_img_object_region_sign), t)
        big_img_object_region_sign = np.array(big_img_object_region_sign)[select_index]
    else:
        big_img_object_region_sign = np.array(big_img_object_region_sign)
    big_img_object_region_sign = np.reshape(big_img_object_region_sign, (-1, 6))
    
    print(big_img_object_region_cell.shape, big_img_object_region_conn.shape, big_img_object_region_sign.shape)

    big_img_object_region = np.concatenate((big_img_object_region_cell, big_img_object_region_conn, big_img_object_region_sign), axis=0)
    
    # while True: # 去掉不同标签的重叠区域
    #     merge_times_sum = 0
    #     # rectangle_1_list = copy.deepcopy(big_img_object_region)
    #     rectangle_1_list = copy.deepcopy(big_img_object_region_sign)
    #     for index, rectangle_1 in enumerate(rectangle_1_list):
    #         for rectangle_2 in rectangle_1_list[index+1:]:
    #             if rectangle_1[0] == rectangle_2[0]:
    #                 continue
    #             # big_img_object_region, merge_once = merge_one_subplot(rectangle_1, rectangle_2, big_img_object_region, t)
    #             big_img_object_region_sign, merge_once = merge_one_subplot(rectangle_1, rectangle_2, big_img_object_region_sign, t)
    #             merge_times_sum = merge_times_sum + merge_once
    #             if merge_once==1:   # 每个区域在每轮合并中仅被合并一次
    #                 break
    #     if merge_times_sum==0:
    #         break

    merge_sign_list, merge_sign_str = merge_sign(big_img_object_region_sign)

    return big_img_object_region, big_img_object_region_cell, big_img_object_region_conn, big_img_object_region_sign, merge_sign_str

def merge_sign(big_img_object_region_sign):
    # 将近邻的字母-数字-符号合并, 返回排序之后的编号集合
    merge_region = {}
    appear_region = {}
    region_index = 0
    big_img_object_region_sign = np.array(big_img_object_region_sign)
    for index, item_0 in enumerate(big_img_object_region_sign):
        # 将近邻的字母-数字-符号合并
        _, center_x_0, center_y_0, width_0, height_0, _ = item_0
        distance = np.sqrt(np.sum(np.power(big_img_object_region_sign[:, 1:3]-item_0[1:3], 2), axis=1)) # 计算每个符号与item_0的距离
        distance[index] = np.max(distance)
        temp = [index] # 自身
        try_times = 0
        while True:
            index_min = np.argmin(distance)
            _, center_x_1, center_y_1, width_1, height_1, _ = big_img_object_region_sign[index_min] # 距离item绝对距离最近的几个目标区域
            try_times += 1
            if try_times==8:
                break
            if (width_0-height_0) * (width_1-height_1)<0: # 文字符号是相同方向
                continue
            if width_0 < height_0: # 水平方向
                if abs(center_y_1-center_y_0) < height_0/4 and abs(center_x_1-center_x_0) < 4*(width_0+width_1)/2:
                    temp.append(index_min)
            else:  # 竖直方向
                if abs(center_x_1-center_x_0) < width_0/4 and abs(center_y_1-center_y_0) < 4*(height_0+height_1)/2:
                    temp.append(index_min)
            distance[index_min] = np.max(distance)

        has_appear = [[index, appear_region[index]] for index in temp if index in appear_region.keys()] # temp内的元素是否在其他区域出现过
        want_merge_regions = list(set([x for _,x in has_appear])) # 将要合并的区域
        merge_region[region_index] = temp
        for index in temp:
            appear_region[index] = region_index
        for will_be_merge_regions in want_merge_regions:
            merge_region[region_index].extend(merge_region[will_be_merge_regions])
            merge_region.pop(will_be_merge_regions)
        for index in merge_region[region_index]:
            appear_region[index] = region_index

        region_index += 1

    result, result_merge = convert_index_2_str(merge_region, big_img_object_region_sign)

    return result, result_merge


def convert_index_2_str(merge_region, big_img_object_region_sign):
    # 将标记的索引转换为标记str
    result = []
    result_merge = []
    label_info = label_get_info(r'data/config.json')
    for region_index in merge_region.keys():
        region_list = list(set(merge_region[region_index]))
        select_regions = big_img_object_region_sign[region_list]
        diff = np.max(select_regions, axis=0) - np.min(select_regions, axis=0)
        if diff[1]>=diff[2]: # 文字为水平方向
            sort_index = np.lexsort((select_regions[:,1],))
            sorted_region = select_regions[sort_index]
            _, x_center, y_center, _, height, probability = np.mean(sorted_region, axis=0)
            width = (sorted_region[-1,1]+sorted_region[-1,3]/2) - (sorted_region[0,1]-sorted_region[0,3]/2)
        else: # 文字为竖直方向
            sort_index = np.lexsort((-select_regions[:,2],))
            sorted_region = select_regions[sort_index]
            _, x_center, y_center, width, _, probability = np.mean(sorted_region, axis=0)
            height = (sorted_region[0,2]+sorted_region[0,4]/2) - (sorted_region[-1,2]-sorted_region[-1,4]/2)
        result.append(sorted_region)
        label_str = ''
        for label in sorted_region:
            label_str += label_info[label[0]]['no'].split('_')[-1]
        result_merge.append([label_str, x_center, y_center, width, height, probability])
    
    return result, result_merge

def save_result(merge_sign_str, file_name):
    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value='{}'.format('Label'))
    ws.cell(row=1, column=2, value='x_min')
    ws.cell(row=1, column=3, value='y_min')
    ws.cell(row=1, column=4, value='x_max')
    ws.cell(row=1, column=5, value='y_max')
    ws.cell(row=1, column=6, value='probability')
    row = 2
    for sign_str, x_center, y_center, width, height, probability in merge_sign_str:
        ws.cell(row=row, column=1, value=sign_str)
        ws.cell(row=row, column=2, value=x_center-width/2)
        ws.cell(row=row, column=3, value=y_center-height/2)
        ws.cell(row=row, column=4, value=x_center+width/2)
        ws.cell(row=row, column=5, value=y_center+height/2)
        ws.cell(row=row, column=6, value=probability)
        row += 1
    wb.save("{}.xlsx".format(file_name))

def show_in_big_img(big_img, big_img_object_region_cell, big_img_object_region_conn, big_img_object_region_sign, \
    merge_sign_str, color_list, draw_sign_str=True, draw_cell=False, draw_conn=False, draw_sign=False):
    # 在大图中将检测到的目标区域框出
    draw = ImageDraw.Draw(big_img)
    if draw_sign_str:
        for sign_str, x_center, y_center, width, height, probability in merge_sign_str:
            draw.rectangle((x_center-width/2, y_center-height/2, x_center+width/2, y_center+height/2),
                        fill=None, outline='red', width=1)
            setFont = ImageFont.truetype('times.ttf', 15)
            draw.text((x_center-width/2, y_center-height/2-15), '{}'.format(sign_str),
                    font=setFont, fill='red', direction=None)

    if draw_cell:
        for label, x_center, y_center, width, height, probability in big_img_object_region_cell:
            draw.rectangle((x_center-width/2, y_center-height/2, x_center+width/2, y_center+height/2),
                        fill=None, outline=color_list[int(label)], width=3)
            setFont = ImageFont.truetype('times.ttf', 15)
            draw.text((x_center-width/2, y_center-height/2-15), '{}'.format(int(label)),
                    font=setFont, fill=color_list[int(label)], direction=None)
    if draw_conn:
        for label, x_center, y_center, width, height, probability in big_img_object_region_conn:
            draw.rectangle((x_center-width/2, y_center-height/2, x_center+width/2, y_center+height/2),
                        fill=None, outline=color_list[int(label)], width=1)
            setFont = ImageFont.truetype('times.ttf', 15)
            draw.text((x_center-width/2, y_center-height/2-15), '{}'.format(int(label)),
                    font=setFont, fill=color_list[int(label)], direction=None)
    if draw_sign:
        for label, x_center, y_center, width, height, probability in big_img_object_region_sign:
            draw.rectangle((x_center-width/2, y_center-height/2, x_center+width/2, y_center+height/2),
                        fill=None, outline=color_list[int(label)], width=1)
            setFont = ImageFont.truetype('times.ttf', 15)
            draw.text((x_center-width/2, y_center-height/2-15), '{}'.format(int(label)),
                    font=setFont, fill=color_list[int(label)], direction=None)


    for file_name, save_data in zip((['元件', '连接点', '标记', '标记-字符']), \
        [big_img_object_region_cell, big_img_object_region_conn, big_img_object_region_sign, merge_sign_str]):
        save_result(save_data, file_name)

    big_img.save('temp.jpg')
    big_img.show()


def label_get_info(config_json_path):
    # 依据config_json_path,返回以label为键的字典
    all_info = {}
    with open(config_json_path, 'r', encoding='utf-8') as json_f:  # 加载组件信息
        component_dict = json.load(json_f)
        for component in component_dict.keys():
            all_info[component_dict[component]['label']] = component_dict[component]
    return all_info

def inference_line(object_region_cell, object_region_conn, config_json_path):
    # 推理得到元件间的连接关系
    json_info = label_get_info(config_json_path)
    result = {}
    for index, cell in enumerate(object_region_cell):
        cell_connection_number = json_info[cell[0]]['connection_number']
        temp_result = {}
        for other_index, other_cell in enumerate(object_region_cell[index+1:]):
            other_cell_connection_number = json_info[other_cell[0]]['connection_number']
            if abs(cell[1]-other_cell[1]) < (cell[3]+other_cell[3])/4:
                x_direction = abs(cell[1]-other_cell[1])
            else:
                x_direction = None
            if abs(cell[2]-other_cell[2]) < (cell[4]+other_cell[4])/4:
                y_direction = abs(cell[2]-other_cell[2])
            else:
                y_direction = None
            temp_result[other_index+index+1] = [x_direction, y_direction]

        for conn in object_region_conn:
            conn_connection_number = json_info[conn_connection_number[0]]['connection_number']
            conn_direction = json_info[conn_connection_number[0]]['direction']

            if abs(cell[1]-conn[1]) < (cell[3]+conn[3])/4:
                x_direction = abs(cell[1]-conn[1])
            else:
                x_direction = None
            if abs(cell[2]-conn[2]) < (cell[4]+conn[4])/4:
                y_direction = abs(cell[2]-conn[2])
            else:
                y_direction = None
            temp_result[other_index+index+1] = [x_direction, y_direction]


if __name__ == '__main__':
    # class_type = 200
    # color_list = ['{:0<7}'.format('#'+str(hex(int(256*256*256/(class_type+2)*x)))[
    #                               2:].upper()) for x in range(class_type)][1:-1]

    # # 分割图片
    # big_pdf_path_list = [r'../data/001_M.pdf']
    # save_small_img_parent_path = r'../data/label5'
    # for big_pdf_path in big_pdf_path_list:
    #     big_img_path_list = pdf_image(pdf_path=big_pdf_path, img_path=save_small_img_parent_path, zoom_x=15, zoom_y=15)
    #     big_img_path_list = [r'../data/006.png']
    #     for big_img_path in big_img_path_list:
    #         big_img, save_small_imgs, mesh_info = mesh_data_test(test_img_path=big_img_path, \
    #             save_small_img_parent_path=save_small_img_parent_path, crop_width=640, crop_height=640, moving_radio=1) # 将大图划分为多个小图

    # # 训练数据
    # mesh_data_train(json_label_path=r'../data/005_0.json', big_img_parent_path=r'../data',
    #                 crop_width=640, crop_height=640, save_img_parent_path=r'../data/train/images/mesh',
    #                 save_label_parent_path=r'../data/train/labels/mesh', color_list=color_list, t=0.85)

    # json转yaml
    # config_yaml(r'data/config.json', r'data/coco128.yaml')

    # 连接关系推理
    big_img_object_region_cell = [[0, 0, 0, 0, 0, 0]]
    big_img_object_region_conn = [[1, 1, 1, 1, 1, 1]]
    config_json_path = r'data/config.json'
    inference_line(big_img_object_region_cell, big_img_object_region_conn, config_json_path)

    # # 测试数据
    # save_img_path = pdf_image(r'../data/test.pdf', '', 2, 2)
    # big_img, save_small_imgs, mesh_info = mesh_data_test(test_img_path=save_img_path[0], crop_width=480, crop_height=480) # 将大图划分为多个小图
    # objects_region = np.random.randint(100, 200, size=(len(mesh_info['top_left_points']), 10, 5))
    # big_img_object_region = merge_objects_region(mesh_info['top_left_points'], objects_region) # 将小图中的目标区域坐标信息，转换合并到对应大图中坐标
    # show_in_big_img(big_img, big_img_object_region, color_list) # 在大图中框选显示目标区域
