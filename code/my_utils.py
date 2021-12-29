from typing import Text
import cv2
import numpy as np
from numpy.core.records import array
import scipy.stats
import copy
import os
import json
import fitz
from sklearn import svm
import random
from PIL import Image
import pickle
import math
import time
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from operator import *
from multiprocessing import Pool, Manager
from gen_coener import gen_all_kinks_conn

min_length_line = 5

class One_Element(object):
    # 定义的元素类（元件、连接点、标记信息）
    def __init__(self, config_path):
        super().__init__()
        self.config_path = config_path
        self.get_config_info()
    
    def one_class(self, label):
        if label not in self.config_info_by_label and label not in self.config_info_by_name:
            print('\033[1;31m 元件配置文件【{}】内未找到标签{} \033[0m'.format(self.config_path, label))  # 有高亮
            label=1000
        if type(label)==int:
            return self.config_info_by_label[label]
        elif type(label)==str:
            return self.config_info_by_name[label]
    
    def get_config_info(self):
        component_dict, all_info =  read_config(self.config_path)

        self.config_info_by_name = component_dict
        self.config_info_by_label = all_info


class One_Object(object):
    # 某类具体目标的类
    def __init__(self, info, object_num):
        super().__init__()
        self.label = info['label']
        self.no = info['no']
        self.type = info['type']
        self.link_direction = info['link_direction']
        self.__link_num = info['link_num']
        self.rectangle = info['rectangle']
        self.has_tab = info['has_tab']
        self.identify_by_tab = info['identify_by_tab']
        self.height_width = info['height_width']
        self.flow_direction = info['flow_direction']
        self.keywords = info['keywords']
        self.show = info['show']
        self.chinese_name = info['chinese_name']
        self.description = info['description']
        self.samples_path = info['samples_path']
        self.__gen_possible_link_direction()
        self.object_num = object_num
        self.link_probability = 1.0
        self.include_cell_sign = [] # 该元件内部包含的元件/拐点
        self.is_inside = [] # 该元件/拐点在某个元件的内部
        
    
    def init_state(self, small_img, position):
        # 依据yolo网络的输出，得到img的实例化的类
        self.x_center = position[1]
        self.y_center = position[2]
        self.width = position[3]
        self.height = position[4]
        self.probability = position[5]
        self.text = position[6]

        if self.type=='cell' or self.type=='conn':
            link_position_four_direction = find_edge_points_one_object(small_img, \
                (self.x_center-self.width/2, self.y_center-self.height/2, self.x_center+self.width/2, self.y_center+self.height/2))
            self.link_num = {}
            self.link_position = {}
            self.link_num_total = 0
            for one_direction, one_direction_link_position in link_position_four_direction.items():
                one_direction_link_num = len(one_direction_link_position)
                self.link_num[one_direction] = one_direction_link_num # 该元件某一方向的可连接点数（定值）
                self.link_position[one_direction] = link_position_four_direction[one_direction] # 该元件某一方向连接点的位置（通过图像分析的方法确定）
                self.link_num_total += one_direction_link_num # 该元件总的可连接点数（定值）
            
            for one_possible in self.possible_direction_num: # 四种可能方向，依次判断
                link_direction, link_num =  one_possible
                match_direction = 0
                for one_direction, one_direction_link_num in zip(link_direction, link_num):
                    if len(link_position_four_direction[one_direction])==one_direction_link_num: # 图像分析确定的连接点数与元件属性中的连接点数匹配成功
                        match_direction += 1
                if match_direction==len(link_direction):
                    self.link_probability=1.0
                    break
                else:
                    self.link_probability=0
            # if self.object_num==1037:
            #     print(self.link_probability, self.link_direction, self.link_num, self.link_position, self.link_num_total)
            self.inference_link_object = [] # 推理时与该元件连接的元件和连接方向
            self.inference_confirm_link_object = [] # 推理时与该元件连接的元件和连接方向,确定的
            self.inference_unsure_link_object = [] # 推理时与该元件连接的元件和连接方向,不确定的
            self.has_confirm_all_link = False # 该元件的所有连接点是否确定
    
    def __gen_possible_link_direction(self):
        # 旋转四个角度，生成不同的可能连接方向
        self.possible_direction_num = []
        for link_direction in [['E','S','W','N'], ['S','W','N','E'], ['W','N','E','S'], ['N','E','S','W']]:
            self.possible_direction_num.append([link_direction, self.__link_num])
    
    def update(self, link_object, link_direction, two_object_link_probability, two_object_link_dis, line_p, rule_num):
        # 每连接一个新的元件之后更新该元件的信息
        # 1、self.inference_link_object
        # if [link_object, link_direction, two_object_link_probability, two_object_link_dis, line_p, rule_num] not in self.inference_link_object:
        if [link_object, link_direction, two_object_link_probability, two_object_link_dis, line_p] not in [x[:-1] for x in self.inference_link_object]:
            self.inference_link_object.append([link_object, link_direction, two_object_link_probability, two_object_link_dis, line_p, rule_num])
            self.has_confirm_all_link_func()
    
    def update_remove(self, link_object):
        # 每连接一个新的元件之后更新该元件的信息
        # 1、self.inference_link_object
        link_object_info = [x for x in self.inference_link_object if link_object in x]
        if len(link_object_info)!=0:
            self.inference_link_object.remove(link_object_info[0])
            self.has_confirm_all_link_func()
    
    def update_confirm_link(self, link_object, rule_num):
        # 确定某一条连接关系, 将相应的信息添加到
        has_update = False
        link_object_info = [x for x in self.inference_link_object if link_object in x] # object与link_object间的连接
        link_object_info = [sorted(link_object_info, key=lambda x: x[2], reverse=True)[0]] # 重合概率越大，可能性越高
        object_info = [x for x in link_object.inference_link_object if self in x] # link_object与object间的连接        
        object_info = [sorted(object_info, key=lambda x: x[2], reverse=True)[0]] # 重合概率越大，可能性越高

        assert len(link_object_info)==1, print(object_info)
        assert len(object_info)==1, print(object_info)

        link_info = link_object_info[0]
        link_info.pop(5)
        link_info.insert(5, rule_num)
        other_link_info = object_info[0]
        other_link_info.pop(5)
        other_link_info.insert(5, rule_num)
        # self_link = [x for x in self.inference_confirm_link_object if link_info[0] in x]
        # other_link = [x for x in link_object.inference_confirm_link_object if other_link_info[0] in x]

        if link_info not in self.inference_confirm_link_object:
            self.inference_confirm_link_object.append(link_info)
            self.has_confirm_all_link_func()
            has_update = True
        if other_link_info not in link_object.inference_confirm_link_object:
            link_object.inference_confirm_link_object.append(other_link_info)
            self.has_confirm_all_link_func()
            has_update = True

        if link_info in self.inference_unsure_link_object: # 删除
            self.update_remove_unsure_link(link_info)
            self.has_confirm_all_link_func()
            has_update = True
        
        return has_update

    def update_unsure_link(self, link_object, rule_num):
        # 添加一条不确定连接关系
        has_update = False
        link_object_info = [x for x in self.inference_link_object if link_object in x] # object与link_object间的连接
        assert len(link_object_info)==1

        link_info = link_object_info[0]
        link_info.pop(5)
        link_info.insert(5, rule_num)

        if link_info not in self.inference_unsure_link_object and link_object_info[0] not in self.inference_confirm_link_object:
            self.inference_unsure_link_object.append(link_object_info[0])
            has_update = True
        
        return has_update

    def update_remove_unsure_link(self, link_object):
        # 删除一条不确定连接关系
        has_update = False
        link_object_info = [x for x in self.inference_unsure_link_object if link_object in x] # object与link_object间的连接
        assert len(link_object_info)==1
        if link_object_info[0] not in self.inference_unsure_link_object:
            self.inference_unsure_link_object.remove(link_object_info[0])
            has_update = True
        
        return has_update

    def update_replace_one_direction(self, one_direction):
        # 依据确定性的连接关系，更新某一方向的连接信息
        has_update = False
        will_remove_object = []
        possible_link_num = copy.deepcopy(len(self.inference_link_object))
        for index in range(possible_link_num):
            _, link_direction, _, _, _, _ = self.inference_link_object[index]
            if link_direction==one_direction:
                if self.inference_link_object[index] not in self.inference_confirm_link_object:
                    will_remove_object.append(self.inference_link_object[index])

        for link_object, _, _, _, _, _ in will_remove_object:
            self.update_remove(link_object) # 删除该元件的无效连接
            self.inference_unsure_link_object = [] # 不确定性关系清空
            has_update = True
        
        self.has_confirm_all_link_func()
        
        return has_update
    
    def has_confirm_all_link_func(self):
        if self.link_probability:
            self.has_confirm_all_link = True
            assert [x for x in ['E', 'S', 'W', 'N'] if x in list(self.link_num)]==[x for x in ['E', 'S', 'W', 'N'] if x in list(self.link_direction)]
            for one_direction in self.link_direction:
                if len([x for x in self.inference_confirm_link_object if x[1]==one_direction])!=self.link_num[one_direction]:
                    self.has_confirm_all_link = False
                    break
    
    def update_object_info(self):
        # 依据确定性连接结果，更新连接信息和元件信息
        has_update_return = False
        has_update = False

        for link_object_info in [x for x in self.inference_link_object]: # 删除彼此不相互连接的元素
            link_object = link_object_info[0]
            if len([x[0] for x in link_object.inference_link_object if self in x])==0:
                self.update_remove(link_object)
                has_update_return = True

        if self.link_probability==1.0: # 元件的连接方向信息确定
            for one_direction in self.link_direction: # 每个确定方向
                if len([x for x in self.inference_confirm_link_object if x[1]==one_direction])==self.link_num[one_direction]:
                    has_update = self.update_replace_one_direction(one_direction)
        
        else: # 不确定连接方向信息元件连接方向信息的确定
            if self.link_num_total==len(self.inference_confirm_link_object): # 不确定连接方向信息元件的可连接点数全部连接完成
                # 是否确定该元件的连接方向信息
                confirm_link_direction_sign = True
                confirm_link_direction = [x[1] for x in self.inference_confirm_link_object]
                for one_direction in self.link_direction: # 每个确定方向
                    if one_direction not in confirm_link_direction:
                        confirm_link_direction_sign = False
                        break
                    if len([x for x in self.inference_confirm_link_object if x[1]==one_direction])!=self.link_num[one_direction]:
                        confirm_link_direction_sign = False
                        break
                if confirm_link_direction_sign:
                    self.link_probability = 1.0 # 确定该元件的方向信息
                    for one_direction in self.link_direction:
                        self.update_replace_one_direction(one_direction) # 更新每个确定方向的连接信息
                    has_update = True

                # 修改元件的连接方向信息
                if not has_update: # 若元件的连接方向未被确定，则判断该元件的连接方向信息是否需要修改
                    for possible_link_direction_num in self.possible_direction_num: # 4个可能方向的逐个判断
                        modify_link_direction_sign = True
                        possible_link_direction, possible_link_num = possible_link_direction_num
                        for possible_link_direction_temp, possible_link_num_temp in zip(possible_link_direction, possible_link_num):
                            if len([x for x in self.inference_confirm_link_object if x[1]==possible_link_direction_temp])!=possible_link_num_temp:
                                modify_link_direction_sign = False
                        if modify_link_direction_sign:
                            self.link_probability = 1.0
                            self.link_direction = [x[0] for x in possible_link_direction_num]
                            link_num_list = [x[1] for x in possible_link_direction_num]
                            self.link_num = {}
                            for one_direction, link_num in zip(self.link_direction, link_num_list):
                                self.link_num[one_direction] = link_num # 更新每个方向的连接点数
                            for one_direction in ['E', 'S', 'W', 'N']:
                                self.update_replace_one_direction(one_direction) # 更新每个确定方向的连接信息
                            has_update = True

        if has_update:
            has_update_return = True
        
        return has_update_return
    
    def estimate_com_link(self, link_object, one_direction):
        # 判断两个元件间满足两者间存在直线连接的前提条件
        # if self.object_num==510 and link_object.object_num==512:
        #     link_object_info = [[x[0].object_num, x[1], x[2], x[3], x[4]] for x in self.inference_link_object if (link_object in x and x[1]==one_direction)]
        #     print(link_object_info)
        #     opposite_link_direction = opposite_direction(one_direction)
        #     object_info = [[x[0].object_num, x[1], x[2], x[3], x[4]] for x in link_object.inference_link_object if (self in x and x[1]==opposite_link_direction)]
        #     print(object_info)
        link_object_info = [x for x in self.inference_link_object if (link_object in x and x[1]==one_direction)]
        if len(link_object_info)==0:
            return False
        else:
            # assert len(link_object_info) == 1
            if len(link_object_info) != 1:
                print(link_object_info)
            opposite_link_direction = opposite_direction(one_direction)
            object_info = [x for x in link_object.inference_link_object if (self in x and x[1]==opposite_link_direction)]
            if len(object_info)==0:
                return False
            else:
                assert len(object_info) == 1
                return True
    
    def estimate_cofirm_link_line_one_direction(self, link_object, one_direction):
        # 评估两个元件之间是否存在确定性的连接关系(直线)
        # 添加规则
        # if self.object_num==510:
        #     print([[x[0].object_num, x[1],x[2],x[3],x[4],x[5]] for x in self.inference_link_object])
        if self.estimate_com_link(link_object, one_direction):
            link_object_info = [x for x in self.inference_link_object if (link_object in x and x[1]==one_direction)][0]
            link_p = link_object_info[2]
            link_dis = link_object_info[3]
            link_line = link_object_info[4]

            # 规则1
            if (link_p>0.98 and link_line>0.98) or (link_dis<=min_length_line and link_p>0.98) or link_line==1: # 两元件间无遮挡，且存在直线
                return True, 1
            # 规则2
            if self.link_probability==1.0: # 主元件的连接方向均确定
                if self.link_num[one_direction] >= len([x for x in self.inference_link_object if x[1]==one_direction]) and link_line>0.8: # 主元件在该方向的连接数小于等于必需连接数
                    return True, 2
            # 规则3
            if link_object.link_probability==1.0: # 副元件的连接方向均确定
                opposite_link_direction = opposite_direction(one_direction)
                if link_object.link_num[opposite_link_direction] >= len([x for x in link_object.inference_link_object if x[1]==opposite_link_direction]) and link_line>0.8: # 副元件在该方向的连接数小于等于必需连接数
                    return True, 3
            # 规则4
            if self.link_probability==1.0 and link_object.link_probability==1.0 and link_p>0.99 and  one_direction in ['E','W'] and 0.9<=link_object.height/self.height<=1.1 and link_line>0.8:
                return True, 4
            # 规则5
            if self.link_probability==1.0 and link_object.link_probability==1.0 and link_p>0.99 and one_direction in ['S','N'] and 0.9<=link_object.width/self.width<=1.1 and link_line>0.8:
                return True, 5
            # 规则6
            confirm_object_info_list = [x for x in self.inference_confirm_link_object if x[1]==one_direction]
            if self.link_probability==1.0 and len(confirm_object_info_list)<self.link_num[one_direction] and link_line>0.8: # 该方向仍存在未连接的接口
                link_object_info_list = [x for x in self.inference_link_object if (link_object not in x and x[1]==one_direction)] # 除副元件之外的其他元件
                good_num = 0
                for other_link_object_info in link_object_info_list:
                    other_link_p = other_link_object_info[2]
                    other_link_dis = other_link_object_info[3]
                    other_link_line = other_link_object_info[4]
                    if (link_p>other_link_p*3 and link_line>other_link_line) or (link_p>other_link_p and link_line>other_link_line*3):
                        good_num += 1
                if good_num==len(link_object_info_list): # 当其他元件与该元件的连接概率和直线长度小于副元件元件的三分之一
                    return True, 6
            # 规则7
            confirm_link_object_info_list = [x for x in link_object.inference_confirm_link_object if x[1]==one_direction]
            opposite_link_direction = opposite_direction(one_direction)
            if link_object.link_probability==1.0 and len(confirm_link_object_info_list)<link_object.link_num[opposite_link_direction] and link_line>0.8: # 该方向仍存在未连接的接口
                link_object_info_list = [x for x in link_object.inference_link_object if (self not in x and x[1]==opposite_link_direction)] # 除主元件之外的其他元件
                good_num = 0
                for other_link_object_info in link_object_info_list:
                    other_link_p = other_link_object_info[2]
                    other_link_dis = other_link_object_info[3]
                    other_link_line = other_link_object_info[4]
                    if (link_p>other_link_p*3 and link_line>other_link_line) or (link_p>other_link_p and link_line>other_link_line*3):
                        good_num += 1
                if good_num==len(link_object_info_list): # 当其他元件与该元件的连接概率和直线长度小于副元件元件的三分之一
                    return True, 7
            # 规则8
            confirm_object_info_list = [x for x in self.inference_confirm_link_object if x[1]==one_direction]
            if self.link_probability==1.0 and len(confirm_object_info_list)<self.link_num[one_direction] and link_line>0.8: # 该方向仍存在未连接的接口
                link_object_info_list = [x for x in self.inference_link_object if (link_object not in x and x[1]==one_direction)] # 除副元件之外的其他元件
                good_num = 0
                for other_link_object_info in link_object_info_list:
                    other_link_p = other_link_object_info[2]
                    other_link_dis = other_link_object_info[3]
                    other_link_line = other_link_object_info[4]
                    if (link_p>max(other_link_p*3, 0.9)) and (link_dis*3<other_link_dis):
                        good_num += 1
                if good_num==len(link_object_info_list): # 当其他元件与该元件的连接概率和直线长度小于副元件元件的三分之一
                    return True, 8
            # 规则9
            confirm_link_object_info_list = [x for x in link_object.inference_confirm_link_object if x[1]==one_direction]
            opposite_link_direction = opposite_direction(one_direction)
            if link_object.link_probability==1.0 and len(confirm_link_object_info_list)<link_object.link_num[opposite_link_direction] and link_line>0.8: # 该方向仍存在未连接的接口
                link_object_info_list = [x for x in link_object.inference_link_object if (self not in x and x[1]==opposite_link_direction)] # 除主元件之外的其他元件
                good_num = 0
                for other_link_object_info in link_object_info_list:
                    other_link_p = other_link_object_info[2]
                    other_link_dis = other_link_object_info[3]
                    other_link_line = other_link_object_info[4]
                    if (link_p>max(other_link_p*3, 0.9)) and (link_dis*3<other_link_dis): # 最大连接概率大于0.9且是其他元件的三倍，距离是其他元件的三分之一
                        good_num += 1
                if good_num==len(link_object_info_list):
                    return True, 9
            
            if True:
                return False, 0

        else:
            return False, 0
    
    def estimate_unsure_link_line_one_direction(self, link_object, one_direction):
        # 评估两个元件之间最可能的连接关系(直线)，不确定性连接
        # 添加规则
        # if self.object_num==510 and link_object.object_num==512:
        #     print([[x[0].object_num, x[1],x[2],x[3],x[4],x[5]] for x in self.inference_link_object], self.estimate_com_link(link_object, one_direction))
        if self.estimate_com_link(link_object, one_direction):
            link_object_info = [x for x in self.inference_link_object if (link_object in x and x[1]==one_direction)][0]
            link_p = link_object_info[2]
            link_dis = link_object_info[3]
            link_line = link_object_info[4]
            # 规则0
            if link_object_info in self.inference_confirm_link_object:
                return False, 0
            # 规则1
            confirm_link_object_info_list = [x for x in self.inference_confirm_link_object if x[1]==one_direction]
            if self.link_probability==1.0 and len(confirm_link_object_info_list)<self.link_num[one_direction] and link_line>0.4: # 该方向仍存在未连接的接口
                still_need_link_num = max(0, self.link_num[one_direction]-len(confirm_link_object_info_list))
                link_object_info_list = [x for x in self.inference_link_object if (link_object not in x and x[1]==one_direction)] # 除主元件之外的其他元件
                good_num = 0
                for other_link_object_info in link_object_info_list:
                    other_link_p = other_link_object_info[2]
                    other_link_dis = other_link_object_info[3]
                    other_link_line = other_link_object_info[4]
                    if (link_p>other_link_p) and ((link_line>other_link_line) or (link_dis<other_link_dis)):
                        good_num += 1
                if good_num>=(len(link_object_info_list)-still_need_link_num+1):
                    return True, 1
            # 规则2
            opposite_link_direction = opposite_direction(one_direction)
            confirm_link_object_info_list = [x for x in link_object.inference_confirm_link_object if x[1]==opposite_link_direction]
            if link_object.link_probability==1.0 and len(confirm_link_object_info_list)<link_object.link_num[opposite_link_direction] and link_line>0.4: # 该方向仍存在未连接的接口
                still_need_link_num = max(0, self.link_num[opposite_link_direction]-len(confirm_link_object_info_list))
                link_object_info_list = [x for x in link_object.inference_link_object if (self not in x and x[1]==opposite_link_direction)] # 除副元件之外的其他元件
                good_num = 0
                for other_link_object_info in link_object_info_list:
                    other_link_p = other_link_object_info[2]
                    other_link_dis = other_link_object_info[3]
                    other_link_line = other_link_object_info[4]
                    if (link_p>other_link_p) and ((link_line>other_link_line) or (link_dis<other_link_dis)):
                        good_num += 1
                if good_num==(len(link_object_info_list)-still_need_link_num+1):
                    return True, 2
            # 规则3
            confirm_link_object_info_list = [x for x in self.inference_confirm_link_object if x[1]==one_direction]
            if link_line>0.9 and len(confirm_link_object_info_list)<self.link_num[one_direction]: # 该方向仍存在未连接的接口
                still_need_link_num = max(0, self.link_num[one_direction]-len(confirm_link_object_info_list))
                link_object_info_list = [x for x in self.inference_link_object if (link_object not in x and x[1]==one_direction)] # 除主元件之外的其他元件
                good_num = 0
                # if self.object_num==510 and link_object.object_num==512:
                #     print(len(link_object_info_list), still_need_link_num, '----------------')
                for other_link_object_info in link_object_info_list:
                    other_link_p = other_link_object_info[2]
                    other_link_dis = other_link_object_info[3]
                    other_link_line = other_link_object_info[4]
                    if (link_p>other_link_p) and ((link_line>other_link_line) or (link_dis<other_link_dis)):
                        good_num += 1
                if good_num>=(len(link_object_info_list)-still_need_link_num+1):
                    return True, 3
            # 规则4
            opposite_link_direction = opposite_direction(one_direction)
            confirm_link_object_info_list = [x for x in link_object.inference_confirm_link_object if x[1]==opposite_link_direction]
            if link_line>0.9 and len(confirm_link_object_info_list)<link_object.link_num[opposite_link_direction]: # 该方向仍存在未连接的接口
                still_need_link_num = max(0, self.link_num[opposite_link_direction]-len(confirm_link_object_info_list))
                link_object_info_list = [x for x in link_object.inference_link_object if (self not in x and x[1]==opposite_link_direction)] # 除副元件之外的其他元件
                good_num = 0
                for other_link_object_info in link_object_info_list:
                    other_link_p = other_link_object_info[2]
                    other_link_dis = other_link_object_info[3]
                    other_link_line = other_link_object_info[4]
                    if (link_p>other_link_p) and ((link_line>other_link_line) or (link_dis<other_link_dis)):
                        good_num += 1
                if good_num==(len(link_object_info_list)-still_need_link_num+1):
                    return True, 4
            
            if True:
                return False, 0

        else:
            return False, 0

    def estimate_confirm_link_line(self, link_object, first=True):
        # 评估两个元件在各个方向上是否存在确定性连接
        has_update = False

        for one_direction in self.link_direction:
            # if self.object_num==479 and link_object.object_num==1391:
            #     # print(self.link_direction)
            #     print([[x[0].object_num, x[1], x[2], x[3], x[4], x[5]] for x in self.inference_link_object], one_direction, '++++++')
            confirm_link, rule_num = self.estimate_cofirm_link_line_one_direction(link_object, one_direction)
            if confirm_link:
                has_update_temp = self.update_confirm_link(link_object, rule_num)
                # if self.object_num==479 and link_object.object_num==1391:
                #     print([[x[0].object_num, x[1], x[2], x[3], x[4], x[5]] for x in self.inference_link_object], one_direction, '++++++')
            else:
                has_update_temp = False
            if has_update_temp:
                has_update = True
    
        return has_update
    
    def estimate_unsure_link_line(self, link_object):
        # 评估两个元件在各个方向上是否存在可能性连接
        has_update = False
        for one_direction in self.link_direction:
            unsure_link, rule_num = self.estimate_unsure_link_line_one_direction(link_object, one_direction)
            if unsure_link:
                has_update_temp = self.update_unsure_link(link_object, rule_num)
            else:
                has_update_temp = False
            if has_update_temp:
                has_update = True
    
        return has_update

def read_config(config_path):
    all_info = {}
    with open(config_path, 'r', encoding='utf-8') as json_f:  # 加载元件属性信息
        component_dict = json.load(json_f)
        for component in component_dict.keys():
            all_info[component_dict[component]['label']] = component_dict[component]
    
    return component_dict, all_info

def opposite_direction(direction):
    # 返回给定方向的反方向
    opposite_dict = {
        'E': 'W',
        'S': 'N',
        'W': 'E',
        'N': 'S',
        'W': 'E',
        'N': 'S',
        'E': 'W',
        'S': 'N',
    }
    return opposite_dict[direction]

def worker(one_object, index, big_img, element, objects_list):
    # 多进程单任务
    label = int(one_object[0])
    one_class_info = element.one_class(label=label)
    one_class = One_Object(one_class_info, index)
    x_center, y_center, width, height = one_object[1], one_object[2], one_object[3], one_object[4]
    img = big_img[int(y_center-height/2-round(min_length_line/2)) : math.ceil(y_center+height/2+round(min_length_line/2)), \
        int(x_center-width/2-round(min_length_line/2)):math.ceil(x_center+width/2+round(min_length_line/2))]
    one_class.init_state(img, one_object)
    objects_list.append(one_class)

    return objects_list


def build_objects_info(all_objects, big_img, config_path):
    # 对检测到的每个目标，为其建立实例化的对象，该对象包含了该目标的所有信息
    # print('\n', all_objects)
    # all_objects = sorted(all_objects_no_sorted, key=itemgetter(1, 2))
    # all_objects = sorted(copy.deepcopy(all_objects), key=lambda x: x[1])
    element = One_Element(config_path)
    objects_list = []
    # managers = Manager()
    # objects_list = managers.list()
    # p = Pool(1)
    for index, one_object in enumerate(all_objects):
        objects_list = worker(one_object, index, big_img, element, objects_list)
        # p.apply_async(worker, (one_object, index, big_img, element, objects_list, temperate_parent_path, ))
    # p.close()
    # p.join()
    
    return objects_list

def img_resize(image1, image2):
    # 等比例缩放图片
    height_1, width_1 = image1.shape[0], image1.shape[1]
    height_2, width_2 = image2.shape[0], image2.shape[1]
    if min(height_1, width_1) > min(height_2, width_2): # 以较小的分辨率为标准
        if (height_2>width_2 and height_1>width_1) or (height_2<width_2 and height_1<width_1):
            height_new, width_new = height_2, width_2
        else:
            height_new, width_new = width_2, height_2
        if width_1 / height_1 >= width_new / height_new:
            image1 = cv2.resize(image1, (width_new, int(height_1 * width_new / width_1)))
        else:
            image1 = cv2.resize(image1, (int(width_1 * height_new / height_1), height_new))
    else:
        if (height_2>width_2 and height_1>width_1) or (height_2<width_2 and height_1<width_1):
            height_new, width_new = height_1, width_1
        else:
            height_new, width_new = width_1, height_1
        if width_2 / height_2 >= width_new / height_new:
            image2 = cv2.resize(image2, (width_new, int(height_2 * width_new / width_2)))
        else:
            image2 = cv2.resize(image2, (int(width_2 * height_new / height_2), height_new))
    return image1, image2

def sift_distance(image1, image2, show=False):
    # 计算两张图片的sift特征匹配点间的距离
    image1, image2 = img_resize(image1, image2)
    # SIFT特征计算
    sift = cv2.xfeatures2d.SIFT_create(nfeatures=4000, contrastThreshold=0.001, edgeThreshold=1000)
    # sift = cv2.xfeatures2d.SURF_create(400)
    psd_kp1, psd_des1 = sift.detectAndCompute(image1, mask=None)
    psd_kp2, psd_des2 = sift.detectAndCompute(image2, mask=None)
    if len(psd_kp1)>=2 and len(psd_kp2)>=2:
        # Flann特征匹配
        FLANN_INDEX_KDTREE = 1
        index_params = dict(algorithm=FLANN_INDEX_KDTREE, trees=5)
        search_params = dict(checks=50)
        flann = cv2.FlannBasedMatcher(index_params, search_params)
        matches = flann.knnMatch(psd_des1, psd_des2, k=2)
        goodMatch = []
        goodMatch_point_1 = []
        goodMatch_point_2 = []
        for m, n in matches:
            # goodMatch是经过筛选的优质配对，如果2个配对中第一匹配的距离小于第二匹配的距离的1/2，基本可以说明这个第一配对是两幅图像中独特的，不重复的特征点,可以保留。
            if m.distance < 0.5*n.distance:
                goodMatch.append(m)
                goodMatch_point_1.append(psd_kp1[m.queryIdx].pt)
                goodMatch_point_2.append(psd_kp2[m.trainIdx].pt)
        goodMatch = np.expand_dims(goodMatch, 1)
        goodMatch_point_1 = np.array(goodMatch_point_1)
        goodMatch_point_2 = np.array(goodMatch_point_2)
        if len(goodMatch_point_1)>0 and len(goodMatch_point_2)>0:
            # goodMatch_point_1[:, 0], goodMatch_point_1[:, 1] = goodMatch_point_1[:, 0]/image1.shape[0], goodMatch_point_1[:, 1]/image1.shape[1]
            # goodMatch_point_2[:, 0], goodMatch_point_2[:, 1] = goodMatch_point_2[:, 0]/image2.shape[0], goodMatch_point_2[:, 1]/image2.shape[1]
            offset = goodMatch_point_1[0, 0] - goodMatch_point_2[0, 1]
            goodMatch_point_2 = goodMatch_point_2 + offset
            dis = np.mean(np.sqrt(np.sum(np.power(goodMatch_point_1 - goodMatch_point_2, 2), axis=1)))
            if show:
                img_out = cv2.drawMatchesKnn(image1, psd_kp1, image2, psd_kp2, goodMatch, None, flags=2)
                ret = cv2.drawKeypoints(image1, psd_kp1, image1)
                cv2.imshow('image', img_out)
                cv2.waitKey(5000)
                cv2.destroyAllWindows()
            return dis
        else:
            return None
    else:
        return None

def calc_hist(gray, bin):
    # 图像在水平和竖直方向的区间灰度统计
    canny_img = cv2.Canny(gray, 80, 150)
    width, height = canny_img.shape[1], canny_img.shape[0]
    width_space, height_space = width/bin, height/bin
    hist_width = []
    hist_height = []
    for index in range(bin):
        width_hist_temp = canny_img[:, round(index*width_space):round((index+1)*width_space)]
        height_hist_temp = canny_img[round(index*height_space):round((index+1)*height_space), :]
        hist_width.append(np.mean(width_hist_temp))
        hist_height.append(np.mean(height_hist_temp))
    hist_width = np.array(hist_width)
    hist_height = np.array(hist_height)

    return hist_width, hist_height

def calc_region(gray, bin=5, canny=True):
    # 图像分区域
    region_list = []
    if canny:
        gray = cv2.Canny(gray, 80, 150)
    width, height = gray.shape[1], gray.shape[0]
    width_space, height_space = width/bin, height/bin
    for width_index in range(bin):
        for  height_index in range(bin):
            region_temp = gray[round(height_index*height_space):round((height_index+1)*height_space), \
                round(width_index*width_space):round((width_index+1)*width_space)]
            region_list.append(region_temp)

    return region_list

def calc_region_hist(gray, bin):
    # 图像分区域灰度统计
    region_hist_list = []
    regions = calc_region(gray, bin)
    for region in regions:
        region_hist_list.append(np.mean(region))
    region_hist_list = np.array(region_hist_list)

    return region_hist_list

def calc_gray_diff(image_gray_1, image_gray_2, bin=5):
    # 计算最小灰度差值
    hist_width_1, hist_height_1 = calc_hist(image_gray_1, bin)
    hist_width_2, hist_height_2 = calc_hist(image_gray_2, bin)
    width_diff = np.min(abs(hist_width_1 - hist_width_2))
    height_diff = np.min(abs(hist_height_1 - hist_height_2))
    
    return width_diff+height_diff

def calc_gray_js(image_gray_1, image_gray_2, bin=5):
    # 计算灰度分布的js散度
    hist_width_1, hist_height_1 = calc_hist(image_gray_1, bin)
    hist_width_2, hist_height_2 = calc_hist(image_gray_2, bin)
    JS_1 = JS_divergence(hist_width_1, hist_width_2)
    JS_2 = JS_divergence(hist_height_1, hist_height_2)

    return JS_1+JS_2

def calc_region_gray_js(image_gray_1, image_gray_2, bin=5):
    # 计算区域灰度分布的js散度
    region_hist_1 = calc_region_hist(image_gray_1, bin)
    region_hist_2 = calc_region_hist(image_gray_2, bin)
    region_JS = JS_divergence(region_hist_1, region_hist_2)

    return region_JS

def calc_region_phash(image_gray_1, image_gray_2, bin=5):
    # 分区域计算感知hash值
    region_list_1 = calc_region(image_gray_1, bin=bin)
    region_list_2 = calc_region(image_gray_2, bin=bin)
    hash_distance_list = []
    for region_1, region_2 in zip(region_list_1, region_list_2):
        hash_1 = phash(region_1)
        hash_2 = phash(region_2)
        hash_distance = hamming_distance(hash_1, hash_2)
        hash_distance_list.append(hash_distance)
    
    return np.mean(np.array(hash_distance_list))

def calc_region_template(image_gray_1, image_gray_2):
    # 分区域模板匹配
    if min(image_gray_2.shape) > min(image_gray_1.shape):
        image_gray_2 = cv2.resize(image_gray_2, (image_gray_1.shape[1], image_gray_1.shape[0]))
    else:
        image_gray_1 = cv2.resize(image_gray_1, (image_gray_2.shape[1], image_gray_2.shape[0]))
    region_list_2 = calc_region(image_gray_2, bin=3)
    region_list_1 = calc_region(image_gray_1, bin=3)
    template_coffe_list = []
    for region_1,region_2 in zip(region_list_1, region_list_2):
        # res = cv2.matchTemplate(region_1, region_2, cv2.TM_SQDIFF_NORMED)
        res = cv2.matchTemplate(image_gray_1, image_gray_2, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
        template_coffe_list.append(max_val)
    
    return np.mean(np.array(template_coffe_list))

def calc_points_distance(image_gray_1, image_gray_2):
    # 计算角点最小距离
    corners_1 = cv2.goodFeaturesToTrack(image_gray_1, 20, 0.01, int(min(image_gray_1.shape)/20))
    corners_1 = np.reshape(corners_1, (-1, 2))
    corners_1[:, 0] = corners_1[:, 0]/image_gray_1.shape[0]
    corners_1[:, 1] = corners_1[:, 1]/image_gray_1.shape[1]
    corners_2 = cv2.goodFeaturesToTrack(image_gray_2, 20, 0.01, int(min(image_gray_2.shape)/20))
    corners_2 = np.reshape(corners_2, (-1, 2))
    corners_2[:, 0] = corners_2[:, 0]/image_gray_2.shape[0]
    corners_2[:, 1] = corners_2[:, 1]/image_gray_2.shape[1]

    if len(corners_1)>len(corners_2):
        base_points = corners_2
        compared_points = corners_1
    else:
        base_points = corners_1
        compared_points = corners_2
    min_distance_list = []
    for point_1 in base_points:
        min_distance_temp = np.min(np.sqrt(np.sum(np.power(compared_points-point_1, 2), axis=1)))
        min_distance_list.append(min_distance_temp)
    
    return np.mean(np.array(min_distance_list))

def calc_region_ssim(image_gray_1, image_gray_2):
    from skimage.metrics import structural_similarity as sk_cpt_ssim
    # 分区域计算感知ssim
    image_gray_2 = cv2.resize(image_gray_2, (image_gray_1.shape[1], image_gray_1.shape[0]))
    SSIM_temp = sk_cpt_ssim(image_gray_2, image_gray_1, multichannel=False)
    
    return -SSIM_temp+1

def JS_divergence(p,q):
    # 计算两个分布的JS散度
    M = (p+q) / 2
    return 0.5*scipy.stats.entropy(p,M)+0.5*scipy.stats.entropy(q, M)

def phash(img):
    # 计算hash值
    img = cv2.resize(img, (8,8), cv2.COLOR_RGB2GRAY)
    avg = sum([sum(img[i]) for i in range(8)])/64
    str = ''
    for i in range(8):
        str += ''.join(map(lambda i: '0' if i< avg else '1', img[i]))
    result = ''
    for i in range(0, 64, 4):
        result += ''.join('%x'%int(str[i: i+4], 2))
    return result
 
def hamming_distance(str1, str2):
    # 计算汉明距离
    if len(str1) != len(str2):
        return
    count = 0
    for i in range(len(str1)):
        if str1[i] != str2[i]:
            count += 1
    return count

def split_link_direction_link_num(direction_num):
    # 分割连接方向和连接点数组成的连接信息
    direction_index = [[x, index] for index,x in enumerate(direction_num) if x in ['E', 'S', 'W', 'N']]
    link_direction = [x for x,y in direction_index]
    link_direction_index = [y for x,y in direction_index]
    link_direction_index.append(len(direction_num))
    link_num_str = [direction_num[link_direction_index[index]+1:link_direction_index[index+1]] for index in range(len(link_direction))]
    link_num = [int(''.join(x)) for x in link_num_str]
    assert len(link_direction)==len(link_num)

    return link_direction, link_num

def direction(result_list, temperate_paths, img1, img2_list, img_index):
    # 计算最小差异（距离）与第二最小差异（距离）的可区分度，返回方向
    link_direction_list = []
    confidence_list = []
    link_num_list = []
    flow_direction_list = []
    img_index_list = []
    for result in result_list:
        if None not in result:
            result_temp = copy.deepcopy(result)
            min_path = temperate_paths[result_temp.index(min(result_temp))]
            min_index = result_temp.index(min(result_temp))
            result_temp[min_index] = max(result_temp)
            second_min_index = result_temp.index(min(result_temp))
            confidence = (result[second_min_index]-result[min_index]) / abs(np.mean(result))
            direction_num = list(os.path.splitext(os.path.split(min_path)[1])[0].split('_')[0])
            link_direction, link_num = split_link_direction_link_num(direction_num)
            flow_direction = list(os.path.splitext(os.path.split(min_path)[1])[0].split('_')[1])

            link_direction_list.append(link_direction)
            confidence_list.append(confidence)
            link_num_list.append(link_num)
            flow_direction_list.append(flow_direction)
            img_index_list.append(min_index)
    link_direction, max_confidence_index, probability  = statistics_direction(link_direction_list, confidence_list)
    
    merge_img(img1, img2_list[img_index_list[max_confidence_index]], probability, img_index)
    if probability!=1:
        pass
        # print(link_direction_list, confidence_list, link_direction_list[max_confidence_index])
        # print("\033[1;31m 元件（{}）方向识别结果可能存在错误！\033[0m".format(img_index))
    possible_link_direction_num = []
    for img_path in temperate_paths:
        direction_num = list(os.path.splitext(os.path.split(img_path)[1])[0].split('_')[0])
        link_directions, link_num = split_link_direction_link_num(direction_num)
        possible_link_direction_num_temp = []
        for possible_direction, possible_num in zip(link_directions, link_num):
            possible_link_direction_num_temp.append([possible_direction, possible_num])
        possible_link_direction_num.append(possible_link_direction_num_temp)

    return link_direction, link_num_list[max_confidence_index], flow_direction_list[max_confidence_index], probability, possible_link_direction_num

def statistics_direction(link_direction_list, confidence_list):
    # 从多种方法计算得到的方向中，通过置信度选取最可能的方向
    result = {}
    for link_direction, confidence in zip(link_direction_list, confidence_list):
        if ''.join(link_direction) in result:
            result[''.join(link_direction)].append(confidence)
        else:
            result[''.join(link_direction)] = [confidence]

    result_max = {}
    for k,v in result.items():
        result[k] = np.mean(np.array(v))
        result_max[k] = np.max(np.array(v))

    max_confidence, max_direction = max(zip(result.values(), result.keys()))
    max_confidence_1, _ = max(zip(result_max.values(), result_max.keys()))
    max_index = link_direction_list.index(list(max_direction))
    if len(result)==1 and (max_confidence>0.3 or max_confidence_1>1):
        probability = 1.0
    else:
        probability = max_confidence

    return list(max_direction), max_index, probability

def svm_class(image_gray_1, image_gray_2_list, temperate_paths):
    height_test_img_features, width_test_img_features = calc_hist(image_gray_1)
    test_img_features = list(height_test_img_features)
    test_img_features.extend(list(width_test_img_features))
    X = []
    Y = []
    for index,image_gray_2 in enumerate(image_gray_2_list):
        for i in range(10):
            height, width = image_gray_2.shape[0], image_gray_2.shape[1]
            height_offset, width_offset = max(2, int(0.05*height)), max(2, int(0.05*width))
            random_height_offset_start, random_width_offset_start = random.randint(0, height_offset), random.randint(0, width_offset)
            random_height_offset_stop, random_width_offset_stop = random.randint(-height_offset, -1), random.randint(-width_offset, -1)
            image_gray_2_temp = image_gray_2[random_height_offset_start: random_height_offset_stop, random_width_offset_start:random_width_offset_stop]
            height_train_img_features, width_train_img_features = calc_hist(image_gray_2_temp)
            train_img_features = list(height_train_img_features)
            train_img_features.extend(list(width_train_img_features))
            X.append(train_img_features)
            Y.append(index)
    clf = svm.SVC(probability=True)
    clf.decision_function_shape = "ovr"
    clf.fit(X, Y)

    return list(1.0-clf.predict_proba([test_img_features])[0])

def merge_img(img_1, img_2, probability, img_index, save_path='./result'):
    # 将两张图片合并在一张图上
    width_height = max(img_1.shape) + max(img_2.shape) + 10
    new_img = np.ones((width_height, width_height))*255.0
    new_img[0:img_1.shape[0], 0:img_1.shape[1]] = img_1
    new_img[(img_1.shape[0]+5):(img_1.shape[0]+5+img_2.shape[0]), (img_1.shape[1]+5):(img_1.shape[1]+5+img_2.shape[1])] = img_2
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    cv2.imwrite('{}/({})_{}.png'.format(save_path, round(probability*100, 2), img_index), new_img)

def get_direction(detect_img_path, temperate_paths, img_type, img_index):
    # 计算得到元件的方向
    sift_result = []
    diff_result = []
    js_result = []
    region_js_result = []
    region_phash = []
    region_template = []
    point_distance = []
    region_ssim = []
    if type(detect_img_path)==str:
        image_gray_1 = cv2.imread(detect_img_path, cv2.IMREAD_GRAYSCALE)
    else:
        image_gray_1 = detect_img_path
    image_gray_2_list = []
    for img_path in temperate_paths:
        image_gray_2 = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
        image_gray_2_list.append(image_gray_2)

        # sift_result.append(sift_distance(image_gray_1, image_gray_2))
        # diff_result.append(calc_gray_diff(image_gray_1, image_gray_2))
        js_result.append(calc_gray_js(image_gray_1, image_gray_2))
        region_js_result.append(calc_region_gray_js(image_gray_1, image_gray_2))
        region_phash.append(calc_region_phash(image_gray_1, image_gray_2))
        region_template.append(calc_region_template(image_gray_1, image_gray_2))
        # point_distance.append(calc_points_distance(image_gray_1, image_gray_2))
        # region_ssim.append(calc_region_ssim(image_gray_1, image_gray_2))
        
    # svm_class_list = svm_class(image_gray_1, image_gray_2_list, temperate_paths)
    # result_list = [js_result, region_js_result, region_template, region_phash]
    result_list = [js_result, region_js_result]
    direction_info = {}
    direction_info['link_direction'], direction_info['link_num'], direction_info['flow_out_direction'], direction_info['probability'],\
         direction_info['possible_link_direction_num']= direction(result_list, temperate_paths, image_gray_1, image_gray_2_list, img_index)

    return direction_info

def read_labelme_json(label_json_path):
    # 读取labelme标记的json文件
    # 返回每个标记框的标签和位置
    label_info = []
    with open(label_json_path, 'r', encoding='utf-8') as f:
        ret_dic = json.load(f)
        for _, item in enumerate(ret_dic['shapes']):
            label = int(item['label'])
            [[x_0, y_0], [x_1, y_1]] = item['points']
            x_0_temp = copy.deepcopy(x_0)
            x_1_temp = copy.deepcopy(x_1)
            y_0_temp = copy.deepcopy(y_0)
            y_1_temp = copy.deepcopy(y_1)
            x_0 = min(x_0_temp, x_1_temp)
            x_1 = max(x_0_temp, x_1_temp)
            y_0 = min(y_0_temp, y_1_temp)
            y_1 = max(y_0_temp, y_1_temp)
            x_center, y_center = (x_0+x_1)/2, (y_0+y_1)/2
            width, height = x_1-x_0, y_1-y_0
            assert x_center > 0 and y_center > 0 and width > 0 and height > 0
            label_info.append([label, x_center, y_center, width, height, 1.0])
    return label_info


def save_label_img(img_path, save_img_path, label_json_path):
    # 将标记的区域依据标签进行保存
    label_info = read_labelme_json(label_json_path)
    big_img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
    for label, x_center, y_center, width, height,_ in label_info:
        temp_img = big_img[round(y_center-height/2):round(y_center+height/2), round(x_center-width/2):round(x_center+width/2)]
        one_class_save_path = os.path.join(save_img_path, str(label))
        if not os.path.exists(one_class_save_path):
            os.makedirs(one_class_save_path)
        _, suffix = os.path.splitext(img_path)
        exists_num = len([x for x in os.listdir(one_class_save_path) if suffix in x])
        cv2.imwrite(os.path.join(one_class_save_path, '{}{}'.format(exists_num, suffix)), temp_img)

def pdf_image(pdf_path, img_path, zoom_x, zoom_y, rotation_angle=0):
    # 将pdf转换为图像
    save_img_path = []
    if not os.path.exists(img_path):
        os.makedirs(img_path)
    with fitz.open(pdf_path) as pdf:
        _, tempfilename = os.path.split(pdf_path)
        filename, _ = os.path.splitext(tempfilename)
        # 逐页读取PDF
        for pg in range(0, pdf.pageCount):
            page = pdf[pg]
            trans = fitz.Matrix(zoom_x, zoom_y).preRotate(rotation_angle)
            pm = page.getPixmap(matrix=trans, alpha=False)
            pm.writePNG(os.path.join(
                img_path, '{}_{}.png'.format(filename, pg)))
            save_img_path.append(os.path.join(
                img_path, '{}_{}.png'.format(filename, pg)))
    return save_img_path

def crop_img_for_labelme(img_path, labelme_path, crop_width, crop_height, save_path):
    # 依据labelme标记的单个元件，从单个元件周围截取一定大小的图片
    big_img = cv2.imread(img_path)
    label_info = read_labelme_json(labelme_path)
    total_height, total_width = big_img.shape[0], big_img.shape[1]
    for position in label_info:
        label, x_center, y_center, width, height, probability = position
        assert crop_height>height, crop_width>width
        crop_img_temp = big_img[max(0, round(y_center-crop_height/2)):min(total_height, round(y_center+crop_height/2)), \
            max(0, round(x_center-crop_width/2)):min(total_width, round(x_center+crop_width/2))]
        
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        exist_num = len([x for x in os.listdir(save_path) if '.png' in x])
        cv2.imwrite(os.path.join(save_path, '{}.png'.format(exist_num)), crop_img_temp)

class img_process():
    # 整图图像处理类
    def __init__(self, object_box, img_path, config_path, show, save_img, save_img_path, save_crop_img):
        img = cv2.imread(img_path)
        file_name = os.path.split(img_path)[1]
        name, suffix = os.path.splitext(file_name)
        self.name = name
        self.suffix = suffix
        self.config_path = config_path

        self.min_length_line = min_length_line

        self.save_img_path = save_img_path # 保存结果图片的位置
        self.show = show # 是否显示图片
        self.save_img = save_img  # 是否保存图片
        self.save_crop_img = save_crop_img # 是否将单个元件的图片保存

        self.img_ori_RGB = copy.deepcopy(img) # RGB图像
        self.img_gray = cv2.cvtColor(self.img_ori_RGB, cv2.COLOR_RGB2GRAY) # 灰度图像
        self.img_bin = self.bin2value(self.img_gray) # 二值化图像

        self.self_all_template = [gen_all_kinks_conn(self.min_length_line*1), \
                    gen_all_kinks_conn(self.min_length_line*3), \
                    gen_all_kinks_conn(self.min_length_line*5)]

        self.object_box = object_box

        self.object_list = build_objects_info(self.object_box, self.img_bin, self.config_path)

    def add_conn_info(self, img_process_for_thin, use_thin_img):
        # 增加拐角类元件信息，该类的主函数
        self.remove_cell_conn_sign() # 移除元件、符号和拐点
        self.thin_line(img_process=img_process_for_thin) # 图像细化
        self.find_corner() # 找角点
        self.find_conn(use_thin_img=use_thin_img)  # 找连接点
        self.object_box.extend(self.conn_info)
        self.object_box = sorted(self.object_box, key=itemgetter(2, 1)) # 排序
        
        self.object_list = build_objects_info(self.object_box, self.img_bin, self.config_path)

        self.draw_edge_points() # 元件与外界的连接点

        self.remove_cell_conn_sign() # 移除元件、符号和拐点

        return self.object_list
    
    def gaussianblur(self, img, kernel_size=5):
        # 高斯模糊
        img_blur = cv2.GaussianBlur(img, (kernel_size, kernel_size), 0)
        return img_blur
    
    def bin2value(self, img, threshold=220, max_value=255):
        # 二值化
        if len(img.shape)==3:
            img = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
        _, img_bin = cv2.threshold(img, threshold, max_value, cv2.THRESH_BINARY)
        return img_bin

    def thin_line(self, img_process):
        # 图像细化
        # img_temp = copy.deepcopy(self.img_gray)
        img_temp = copy.deepcopy(self.img_no_sign_zero)

        if img_process:
            img_blur = self.gaussianblur(img_temp)
            img_bin = ~self.bin2value(img_blur)
            kernel = np.ones((self.min_length_line, self.min_length_line), np.uint8)
            img_bin = cv2.dilate(img_bin, kernel, iterations=1)
            img_bin = cv2.erode(img_bin, kernel, iterations=1)
        else:
            img_bin = ~self.bin2value(img_temp)

        img_thinning = cv2.ximgproc.thinning(img_bin, thinningType=cv2.ximgproc.THINNING_ZHANGSUEN)
        # from skimage import morphology
        # img_bin[img_bin==255] = 1
        # skeleton0 = morphology.skeletonize(img_bin)
        # img_thinning = skeleton0.astype(np.uint8)*255

        img_thinning = ~img_thinning

        if self.save_img:
            cv2.imwrite(r'{}\thin_{}{}'.format(self.save_img_path, self.name, self.suffix), img_thinning)
        if self.show:
            img_thinning = Image.fromarray(img_thinning)
            img_thinning.show()
            img_thinning = np.array(img_thinning)

        self.img_thinning = img_thinning

    def find_corner(self):
        # 查找角点
        # fast = cv2.FastFeatureDetector_create()
        # kps = fast.detect(self.img_thinning, None)
        # points = [[list(x.pt)] for x in kps]
        # points = cv2.goodFeaturesToTrack(self.img_thinning, 10000, 0.0001, int(self.min_length_line*1), useHarrisDetector=False)
        points = cv2.goodFeaturesToTrack(self.img_thinning, 1000000, 0.001, int(self.min_length_line*1), useHarrisDetector=False)
        # stop_criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 30, 0.001)
        # points = cv2.cornerSubPix(self.img_thinning, points, (5, 5), (-1, -1), stop_criteria)
        corner_list = []
        margin_pix = int(1.5*self.min_length_line)
        object_box_array = np.array([x[:6] for x in self.object_box])
        
        for point in points:
            # 对角点的位置进行修正，使其位于中心
            point_x, point_y = point[0][0], point[0][1]

            # # 当某角点周围不存在直线段时，删除该角点
            # crop_img_bigger = self.img_thinning[int(point_y-2*margin_pix):math.ceil(point_y+2*margin_pix),\
            #                 int(point_x-2*margin_pix):math.ceil(point_x+2*margin_pix)]
            # row_has_line = [[index for index in range(len(one_row_data)-self.min_length_line) if np.mean(one_row_data[index:(index+self.min_length_line)])==0] for one_row_data in crop_img_bigger]
            # crop_img_bigger_T = crop_img_bigger.T
            # column_has_line = [[index for index in range(len(one_column_data)-self.min_length_line) if np.mean(one_column_data[index:(index+self.min_length_line)])==0] for one_column_data in crop_img_bigger_T]
            # if set([len(x) for x in row_has_line])==set([0]) and set([len(x) for x in column_has_line])==set([0]):
            #     continue

            crop_img = self.img_thinning[int(point_y-margin_pix):math.ceil(point_y+margin_pix+1),\
                                        int(point_x-margin_pix):math.ceil(point_x+margin_pix+1)]
            crop_img_bin = self.img_bin[int(point_y-margin_pix):math.ceil(point_y+margin_pix+1),\
                                        int(point_x-margin_pix):math.ceil(point_x+margin_pix+1)]

            if min(crop_img.shape)>0:
                # 对角点位置进行修正，使其位于中心，分单接口和多接口两种情况
                height_mean = np.mean(crop_img, axis=0)
                width_mean = np.mean(crop_img, axis=1)
                width_center = np.mean(np.where(height_mean==min(height_mean))[0])
                height_center = np.mean(np.where(width_mean==min(width_mean))[0])

                if min(height_mean) > 255*(1-3/(2*margin_pix)) and abs(margin_pix-width_center)<=1.5*self.min_length_line: # 框内只有一条直线的情况（角点只有一个接口，水平）
                    height_mean_bin = np.mean(crop_img_bin, axis=0)
                    width_center_bin = np.mean(np.where( abs(np.diff(height_mean_bin))==max(abs(np.diff(height_mean_bin))) )[0])
                    # point_x = point_x - margin_pix + width_center_bin
                    # print(crop_img_bin, width_center_bin)
                else:
                    height_mean_bin = np.mean(crop_img_bin, axis=0)
                    width_center_bin = np.mean(np.where(height_mean_bin==min(height_mean_bin))[0])
                    # print(crop_img_bin, width_center_bin)
                point_x = point_x - margin_pix + width_center_bin

                if min(width_mean) > 255*(1-3/(2*margin_pix)) and abs(margin_pix-height_center)<=1.5*self.min_length_line: # 框内只有一条直线的情况（角点只有一个接口，竖直）
                    width_mean_bin = np.mean(crop_img_bin, axis=1)
                    height_center_bin = np.mean(np.where( abs(np.diff(width_mean_bin))==max(abs(np.diff(width_mean_bin))) )[0])
                    # point_y = point_y - margin_pix + height_center_bin
                else:
                    width_mean_bin = np.mean(crop_img_bin, axis=1)
                    height_center_bin = np.mean(np.where(width_mean_bin==min(width_mean_bin))[0])
                    # print(crop_img_bin, height_center_bin)
                point_y = point_y - margin_pix + height_center_bin

                # 删除不符合要求的角点
                # 拐点周围应该存在直线，当不存在直线时删除该角点
                point_x_int = round(point_x)
                point_y_int = round(point_y)
                if (np.mean(self.img_bin[point_y_int, (point_x_int+1):(point_x_int+2*self.min_length_line+1)])==0 or \
                    np.mean(self.img_bin[point_y_int, (point_x_int-2*self.min_length_line):point_x_int])==0) or \
                    (np.mean(self.img_bin[(point_y_int+1):(point_y_int+2*self.min_length_line+1), point_x_int])==0 or \
                    np.mean(self.img_bin[(point_y_int-2*self.min_length_line):point_y_int, point_x_int])==0):
                    # 删除yolo检测的元件标记框内的角点
                    if min(object_box_array.shape)!=0:
                        min_x_y_space = np.array(abs(object_box_array[:, 1:3] - [point_x, point_y]) - object_box_array[:, 3:5]/2)
                        min_xy_space = np.max(min_x_y_space, axis=1)
                    else:
                        min_xy_space = [1]
                    if min(min_xy_space)>0 and [round(point_x), round(point_y)] not in corner_list:
                        corner_list.append([round(point_x), round(point_y)])

        if self.show or self.save_img:
            img_temp = copy.deepcopy(self.img_ori_RGB)
            # img_temp = copy.deepcopy(self.img_thinning)
            for corner in corner_list:
                cv2.circle(img_temp, (corner[0], corner[1]), 2, (255,0,0), 2)
            if self.save_img:
                cv2.imwrite(r'{}\corner_{}{}'.format(self.save_img_path, self.name, self.suffix), img_temp)
            if self.show:
                img_temp = Image.fromarray(img_temp)
                img_temp.show()

        self.corner_list = corner_list
    
    def find_conn(self, use_thin_img):
        # 拐角识别
        # 从config文件中读取每类拐角对应的label
        conn_info_label_direction = {}
        _, all_info = read_config(self.config_path)
        for label in all_info:
            if all_info[label]['type']=='conn':
                link_direction = [direction for direction, num in zip(all_info[label]['link_direction'], all_info[label]['link_num']) if num>0]
                sorted_link_direction = [x for x in ['E', 'S', 'W', 'N'] if x in link_direction]
                conn_info_label_direction[''.join(sorted_link_direction)] = label

        conn_info = []
        conn_info_temp = []
        for index, corner_x_y in enumerate(self.corner_list):
            x_center, y_center = corner_x_y
            small_img_list = []
            for min_space in range(int(self.min_length_line*0.5), int(self.min_length_line*2), 3):
                x_min = x_center-min_space
                y_min = y_center-min_space
                x_max = x_center+min_space
                y_max = y_center+min_space
                if use_thin_img: # 使用细化的图像判断拐角
                    small_img = self.img_thinning[y_min:(y_max+1), x_min:(x_max+1)]
                else:
                    small_img = self.img_bin[y_min:(y_max+1), x_min:(x_max+1)]
                small_img_ori = self.img_bin[y_min:(y_max+1), x_min:(x_max+1)]
                small_img_list.append([small_img, small_img_ori, min_space])

            link_direction, dis, min_space = self.estimate_conn(small_img_list)
            # link_direction, dis, min_space = self.estimate_conn_v2(index, small_img_list)
            if dis!=None:
                sorted_link_direction = [x for x in ['E', 'S', 'W', 'N'] if x in link_direction]
                label = conn_info_label_direction[''.join(sorted_link_direction)] # 拐点的label
                assert label>=1001
                conn_info.append([label, x_center, y_center, min_space*2, min_space*2, dis, ''])
                conn_info_temp.append([x_center-min_space, y_center-min_space, x_center+min_space, y_center+min_space])
        
        # 当一个拐点在另一个拐点内部时，删除其中一个拐点
        will_delete_conn = []
        conn_info_temp = np.array(conn_info_temp)
        for conn in conn_info:
            width_min, height_min, width_max, height_max = conn[1]-conn[3]/2, conn[2]-conn[4]/2, conn[1]+conn[3]/2, conn[2]+conn[4]/2
            dis = conn_info_temp - [width_min, height_min, width_max, height_max]
            more_big_box = [x for x in dis if (x[0]<=(0.05*conn[3]) and x[1]<=(0.05*conn[4]) and x[2]>=-(0.05*conn[3]) and x[3]>=-(0.05*conn[4]))]
            if len(more_big_box)>1: # 去除自身，因此要大于1，自身与自身的差值为[0,0,0,0]
                will_delete_conn.append(conn)
        for conn in will_delete_conn:
            conn_info.remove(conn)

        self.conn_info = conn_info

        if self.show or self.save_img and False:
            # img_temp = copy.deepcopy(self.img_ori_RGB)
            img_temp = copy.deepcopy(self.img_thinning)
            for label, x_center, y_center, width, height, p, text in conn_info:
                rectangle_color = (255, 0, 0)
                cv2.rectangle(img_temp, (int(x_center-width/2), int(y_center-height/2)), (int(x_center+width/2), int(y_center+height/2)), rectangle_color, 2)
            if self.save_img:
                cv2.imwrite(r'{}\conn_{}{}'.format(self.save_img_path, self.name, self.suffix), img_temp)
            if self.show:
                img_temp = Image.fromarray(img_temp)
                img_temp.show()
                img_temp = np.array(img_temp)

    def remove_cell_conn_sign(self):
        # 将图中的元件\拐点\文字区域清除【置为(255,255,255)】
        cell_sign_pos_list = []
        img_temp_cell = copy.deepcopy(self.img_gray)
        img_temp_sign = copy.deepcopy(self.img_gray)
        img_temp_sign_zero = copy.deepcopy(self.img_gray)
        img_temp_cell_sign = copy.deepcopy(self.img_gray)
        img_temp_cell_conn = copy.deepcopy(self.img_gray)
        img_temp_cell_conn_sign = copy.deepcopy(self.img_gray)
        for object in self.object_list:
            if object.type in ['sign' ,'cell', 'conn']:
                start_x = int(object.x_center-object.width/2)
                start_y = int(object.y_center-object.height/2)
                stop_x = math.ceil(object.x_center+object.width/2)
                stop_y = math.ceil(object.y_center+object.height/2)
                small_img = img_temp_cell_sign[start_y:stop_y, start_x:stop_x]
                if self.save_crop_img:
                    save_small_img_parent_path = r'{}/crop_img/{}'.format(self.save_img_path, self.name, object.label)
                    if not os.path.exists(save_small_img_parent_path):
                        os.makedirs(save_small_img_parent_path)
                    save_small_img_path = os.path.join(save_small_img_parent_path, '{}_{}{}'.format(object.label, object.object_num, self.suffix))
                    cv2.imwrite(save_small_img_path, small_img)
                
                # 截取区域超过图片大小
                white_img = np.ones((min(stop_y, self.img_gray.shape[0])-start_y, min(stop_x, self.img_gray.shape[1])-start_x))*255
                
                img_temp_cell_conn_sign[start_y:stop_y, start_x:stop_x] = white_img
                if object.type=='cell':
                    img_temp_cell[start_y:stop_y, start_x:stop_x] = white_img
                if object.type=='sign':
                    img_temp_sign[start_y:stop_y, start_x:stop_x] = white_img
                    margin_pix = 2
                    W_LINE = np.min(img_temp_sign_zero[start_y:stop_y, max(0,start_x-margin_pix)])
                    E_LINE = np.min(img_temp_sign_zero[start_y:stop_y, min(stop_x+margin_pix, img_temp_sign.shape[1])])
                    N_LINE = np.min(img_temp_sign_zero[max(0,start_y-margin_pix), start_x:stop_x])
                    S_LINE = np.min(img_temp_sign_zero[min(stop_y+margin_pix, img_temp_sign.shape[0]), start_x:stop_x])
                    if W_LINE>=225.0 and E_LINE>=225.0 and N_LINE>=225.0 and S_LINE>=225.0:
                        img_temp_sign_zero[start_y:stop_y, start_x:stop_x] = white_img
                if object.type in ['cell', 'conn']:
                    img_temp_cell_conn[start_y:stop_y, start_x:stop_x] = white_img
                if object.type in ['cell', 'sign']:
                    img_temp_cell_sign[start_y:stop_y, start_x:stop_x] = white_img
                cell_sign_pos_list.append([start_x,start_y, stop_x,stop_y])
        if self.save_img:
            cv2.imwrite(r'{}\remove_cell_{}{}'.format(self.save_img_path, self.name, self.suffix), img_temp_cell)
            cv2.imwrite(r'{}\remove_sign_{}{}'.format(self.save_img_path, self.name, self.suffix), img_temp_sign)
            cv2.imwrite(r'{}\remove_sign_zero_{}{}'.format(self.save_img_path, self.name, self.suffix), img_temp_sign_zero)
            cv2.imwrite(r'{}\remove_cell_sign_{}{}'.format(self.save_img_path, self.name, self.suffix), img_temp_cell_sign)
            cv2.imwrite(r'{}\remove_cell_conn_{}{}'.format(self.save_img_path, self.name, self.suffix), img_temp_cell_conn)
            cv2.imwrite(r'{}\remove_cell_conn_sign_{}{}'.format(self.save_img_path, self.name, self.suffix), img_temp_cell_conn_sign)
        if self.show:
            img_temp = Image.fromarray(img_temp_cell_sign)
            img_temp.show()
            img_temp = np.array(img_temp)

        self.img_no_cell_conn_sign = img_temp_cell_conn_sign
        self.img_no_cell_sign = img_temp_cell_sign
        self.img_no_cell_conn = img_temp_cell_conn
        self.img_no_cell = img_temp_cell
        self.img_no_sign = img_temp_sign
        self.img_no_sign_zero = img_temp_sign_zero
        self.cell_sign_pos = cell_sign_pos_list


    def draw_edge_points(self):
        # 绘制元件与外界连接点
        img_temp = copy.deepcopy(self.img_ori_RGB)
        if self.show or self.save_img_path and False:
            for object in self.object_list:
                if object.type in ['cell', 'conn']:
                    for one_direction in ['E', 'S', 'W', 'N']:
                        corner = object.link_position[one_direction]
                        for points in corner:
                            cv2.line(img_temp, (int(points[0][0]), int(points[0][1])), (int(points[1][0]), int(points[1][1])), (255,0,0), 2)
                    if object.type=='cell':
                        rectangle_color = (0, 255, 0)
                    elif object.type=='conn':
                        rectangle_color = (0, 0, 255)
                    cv2.rectangle(img_temp, (int(object.x_center-object.width/2), int(object.y_center-object.height/2)), (int(object.x_center+object.width/2), int(object.y_center+object.height/2)), rectangle_color, 1)
            if self.save_img:
                cv2.imwrite(r'{}\edge_{}{}'.format(self.save_img_path, self.name, self.suffix), img_temp)
            if self.show:
                img_temp = Image.fromarray(img_temp)
                img_temp.show()
                img_temp = np.array(img_temp)

    def estimate_conn(self, small_img_list):
        # 评估小图是否为拐点，返回连接方向
        def func_one(crop_img, crop_img_ori, link_direction):
            # 找到拐角中心，并判断从中心出来的线是否为连续的
            height_mean = np.mean(crop_img, axis=0)
            width_mean = np.mean(crop_img, axis=1)
            width_center = int(np.mean(np.where(height_mean==min(height_mean))[0]))
            height_center = int(np.mean(np.where(width_mean==min(width_mean))[0]))
            # print(crop_img, width_center, height_center)
            for one_direction in link_direction:
                if one_direction=='E':
                    height_line = np.mean(crop_img_ori[:, width_center:], axis=0)
                    sorted_value = sorted(height_line, reverse=True)
                elif one_direction=='W':
                    height_line = np.mean(crop_img_ori[:, :width_center], axis=0)
                    sorted_value = sorted(height_line, reverse=True)
                elif one_direction=='S':
                    width_line = np.mean(crop_img_ori[height_center:, :], axis=1)
                    sorted_value = sorted(width_line, reverse=True)
                elif one_direction=='N':
                    width_line = np.mean(crop_img_ori[:height_center, :], axis=1)
                    sorted_value = sorted(width_line, reverse=True)
                if len(sorted_value)>=1:
                    if sorted_value[0]==255:
                        return False
            else:
                return True

        if np.mean(small_img_list[0][0])==255: # 当元件角点周围是空白时
            link_direction=[]
            return link_direction, None, None
        
        # # 当某角点周围不存在直线段时，删除该角点(直线段长度为2*self.min_length_line)
        # crop_img_bigger = small_img_list[1][0]
        # row_has_line = [[index for index in range(len(one_row_data)-self.min_length_line) if np.mean(one_row_data[index:(index+self.min_length_line)])==0] for one_row_data in crop_img_bigger]
        # crop_img_bigger_T = crop_img_bigger.T
        # column_has_line = [[index for index in range(len(one_column_data)-self.min_length_line) if np.mean(one_column_data[index:(index+self.min_length_line)])==0] for one_column_data in crop_img_bigger_T]
        # if set([len(x) for x in row_has_line])==set([0]) and set([len(x) for x in column_has_line])==set([0]):
        #     link_direction=[]
        #     return link_direction, None, None

        height, width = small_img_list[0][0].shape[0], small_img_list[0][0].shape[1]
        if min(height, width)!=0:
            for small_img, small_img_ori, min_space in small_img_list:
                bin = 4
                pix = small_img.shape[0]/bin
                # region_list = calc_region(small_img, bin=4, canny=False)
                four_direction_region = {}
                four_direction_region['E'] = small_img[int(1*pix):int(3*pix), int(3*pix):]
                four_direction_region['W_N'] = small_img[:math.ceil(1*pix), :math.ceil(1*pix)]
                four_direction_region['S'] = small_img[int(3*pix):, int(1*pix):math.ceil(3*pix)]
                four_direction_region['E_N'] = small_img[:math.ceil(pix), math.ceil(3*pix):]
                four_direction_region['W'] = small_img[int(1*pix):int(3*pix), :math.ceil(1*pix)]
                four_direction_region['E_S'] = small_img[int(3*pix):, int(3*pix):]
                four_direction_region['N'] = small_img[:int(1*pix), int(1*pix):math.ceil(3*pix)]
                four_direction_region['W_S'] = small_img[:math.ceil(1*pix), int(3*pix):]

                # region_list = calc_region(small_img, bin=3, canny=False)
                # four_direction_region = {}
                # four_direction_region['E'] = region_list[5]
                # four_direction_region['W_N'] = region_list[0]
                # four_direction_region['S'] = region_list[7]
                # four_direction_region['E_N'] = region_list[2]
                # four_direction_region['W'] = region_list[3]
                # four_direction_region['E_S'] = region_list[8]
                # four_direction_region['N'] = region_list[1]
                # four_direction_region['W_S'] = region_list[6]
                
                link_direction = []
                for one_direction in ['E', 'S', 'W', 'N']:
                    region_data = four_direction_region[one_direction]/255
                    if min(region_data.shape)>0:
                        if one_direction in ['E', 'W']:
                            line_p = 1 - min(np.mean(region_data, axis=1))
                            white_line = max(np.mean(region_data, axis=1))
                        if one_direction in ['S', 'N']:
                            line_p = 1 - min(np.mean(region_data, axis=0))
                            white_line = max(np.mean(region_data, axis=0))
                        if line_p == 1 and white_line == 1:
                            link_direction.append(one_direction)

                if set(link_direction)!=set(['E', 'W']) and set(link_direction)!=set(['S', 'N']) and len(link_direction)>=2:
                    good_corner_num = 0
                    for corner_direction in ['W_N', 'E_N', 'W_S', 'E_S']:
                        if np.min(four_direction_region[corner_direction])==255:
                            good_corner_num += 1
                    if good_corner_num>=4:
                        if func_one(small_img, small_img_ori, link_direction):
                        # if True:
                            return link_direction, 1.0, min_space
                        else:
                            link_direction=[]
                            return link_direction, None, None
            else:
                return link_direction, None, min_space
        
        else:
            link_direction=[]
            return link_direction, None, None
    
    def estimate_conn_v2(self, img_index, small_img_list):
        # 评估小图是否为拐点，返回连接方向
        height, width = small_img_list[0][0].shape[0], small_img_list[0][0].shape[1]
        if min(height, width)!=0:
            for index, (small_img, small_img_ori, min_space) in enumerate(small_img_list):
                tempalte_dict = self.self_all_template[index]
                score, direction = template_img(small_img, tempalte_dict)
                if score>=0.8:
                    cv2.imwrite(r'11/{}_{}_{}.jpg'.format(score*100, direction, img_index), small_img)
                    return list(direction), 1.0, min_space
            link_direction=[]
            return link_direction, None, None
        else:
            link_direction=[]
            return link_direction, None, None


def template_img(img, tempalte_dict):
    template_coffe_list = []
    template_direction_list = []
    for item in tempalte_dict:
        res = cv2.matchTemplate(img, tempalte_dict[item], cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
        template_coffe_list.append(max_val)
        template_direction_list.append(item)
    index = template_coffe_list.index(max(template_coffe_list))
    direction = template_direction_list[index]
    score = template_coffe_list[index]

    return score, direction
    
def narrow_cell_posi_img(img_path, object_box):
    # 缩小元件框, 根据最大连通域缩小
    img_gray = cv2.imread(img_path, 0)
    _, img_bin = cv2.threshold(img_gray, 220, 255, cv2.THRESH_BINARY_INV)
    new_object_box = []
    for one_object_box in object_box:
        label = one_object_box[0]
        x_center = one_object_box[1]
        y_center = one_object_box[2]
        width = one_object_box[3]
        height = one_object_box[4]
        probability = one_object_box[5]
        small_img_bin = img_bin[int(y_center-height/2):int(y_center+height/2), int(x_center-width/2):int(x_center+width/2)]
        contours, hierarchy = cv2.findContours(small_img_bin, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

        bounding_boxes = [[cv2.contourArea(cnt), cv2.boundingRect(cnt)] for cnt in contours]
        if label!=0:
            x, y, w, h = sorted(bounding_boxes, key=lambda x: x[0], reverse=True)[0][1] # 元件，最大的矩形
            if w<=width*0.5 or h<=height*0.5: # 当最大外界矩形的长or宽小于原图片的一半，则保持不变
                x, y, w, h = 0, 0, width, height
        else:
            bounding_boxes_array = np.array(bounding_boxes)[:, 1] # 字母数字等符号，所有矩形的最大外接矩形
            bounding_boxes_array = np.array([[x, y, x+w, y+h] for x,y,w,h in bounding_boxes_array])
            x_min, y_min, x_max, y_max = min(bounding_boxes_array[:, 0]), min(bounding_boxes_array[:, 1]), max(bounding_boxes_array[:, 2]), max(bounding_boxes_array[:, 3])
            x, y, w, h = x_min, y_min, x_max-x_min, y_max-y_min
        
        # small_img_bin = ~small_img_bin
        # small_img_bin = cv2.cvtColor(small_img_bin, cv2.COLOR_GRAY2RGB)
        # cv2.rectangle(small_img_bin, (x,y), (x+w, y+h), (255,0,0), 1)
        # cv2.imshow('111', small_img_bin)
        # cv2.waitKey(5000)
        
        narrow_box = [label, int(x+x_center-width/2+w/2), int(y+y_center-height/2+h/2), math.ceil(w), math.ceil(h), probability]
        new_object_box.append(narrow_box)
    
    return new_object_box

def adjust_cell_posi_img(img_path, object_box):
    # 调整标记框的位置
    img_gray = cv2.imread(img_path, 0)
    _, img_bin = cv2.threshold(img_gray, 220, 255, cv2.THRESH_BINARY_INV)
    new_object_box = []
    for one_object_box in object_box:
        label = one_object_box[0]
        x_center = one_object_box[1]
        y_center = one_object_box[2]
        width = one_object_box[3]
        height = one_object_box[4]
        probability = one_object_box[5]
        margin_pix_x = 2*min_length_line
        margin_pix_y = 2*min_length_line
        small_img_bin = img_bin[int(y_center-height/2-margin_pix_y):int(y_center+height/2+margin_pix_y),\
                        int(x_center-width/2-margin_pix_x):int(x_center+width/2+margin_pix_x)]
        
        small_img_bin_mean_x = np.mean(small_img_bin, axis=0)
        small_img_bin_mean_y = np.mean(small_img_bin, axis=1)
        
        diff_list_x = np.diff(small_img_bin_mean_x)[min_length_line:]
        std_list_x = [np.std(small_img_bin_mean_x[:index]) for index in range(min_length_line, len(small_img_bin_mean_x)-1)]
        sudden_points_w = [index+min_length_line for index, (x,y) in enumerate(zip(diff_list_x, std_list_x)) if abs(x)>=max(3*y,3)]

        trans_small_img_bin_mean_x = np.flip(small_img_bin_mean_x)
        diff_list_x = np.diff(trans_small_img_bin_mean_x)[min_length_line:]
        std_list_x = [np.std(trans_small_img_bin_mean_x[:index]) for index in range(min_length_line, len(trans_small_img_bin_mean_x)-1)]
        sudden_points_e = [(index+min_length_line)-1 for index, (x,y) in enumerate(zip(diff_list_x, std_list_x)) if abs(x)>=max(3*y,3)]
        
        diff_list_y = np.diff(small_img_bin_mean_y)[min_length_line:]
        std_list_y = [np.std(small_img_bin_mean_y[:index]) for index in range(min_length_line, len(small_img_bin_mean_y)-1)]
        sudden_points_n = [index+min_length_line for index, (x,y) in enumerate(zip(diff_list_y, std_list_y)) if abs(x)>=max(3*y,3)]
    
        trans_small_img_bin_mean_y = np.flip(small_img_bin_mean_y)
        diff_list_y = np.diff(trans_small_img_bin_mean_y)[min_length_line:]
        std_list_y = [np.std(trans_small_img_bin_mean_y[:index]) for index in range(min_length_line, len(trans_small_img_bin_mean_y)-1)]
        sudden_points_s = [(index+min_length_line)-1 for index, (x,y) in enumerate(zip(diff_list_y, std_list_y)) if abs(x)>=max(3*y,3)]

        # print(small_img_bin_mean_x)
        # print(sudden_points_w[:1], sudden_points_e[:1])
        # print(small_img_bin_mean_y)
        # print(sudden_points_n[:1], sudden_points_s[:1])
        if len(sudden_points_w)>0:
            x_start = x_center - width/2 -margin_pix_x + sudden_points_w[0]
            if 0.25*width<(x_center-x_start)<0.75*width:
                # print(x_start - (x_center - width/2))
                x_start = x_start
            else:
                x_start = x_center - width/2
        else:
            x_start = x_center - width/2

        if len(sudden_points_e)>0:
            x_stop = x_center + width/2 + margin_pix_x - sudden_points_e[0]
            if 0.25*width>(x_stop-x_center)>0.75*width:
                # print('22222222222')
                x_stop = x_stop
            else:
                x_stop = x_center + width/2
        else:
            x_stop = x_center + width/2

        if len(sudden_points_n)>0:
            y_start = y_center - height/2 - margin_pix_y + sudden_points_n[0]
            if 0.25*height>(y_center-y_start)>0.75*height:
                print('3333333333')
                y_start = y_start
            else:
                y_start = y_center - height/2
        else:
            y_start = y_center - height/2

        if len(sudden_points_s)>0:
            y_stop = y_center + height/2 + margin_pix_y - sudden_points_s[0]
            if 0.25*height>(x_stop-y_center)>0.75*height:
                print('444444444444444444')
                y_stop = y_stop
            else:
                y_stop = y_center + height/2
        else:
            y_stop = y_center + height/2
        
        # small_img_bin = ~small_img_bin
        # small_img_bin = cv2.cvtColor(small_img_bin, cv2.COLOR_GRAY2RGB)
        # cv2.rectangle(small_img_bin, (x,y), (x+w, y+h), (255,0,0), 1)
        # cv2.imshow('111', small_img_bin)
        # cv2.waitKey(5000)
        
        narrow_box = [label, int((x_start+x_stop)/2), int((y_start+y_stop)/2), math.ceil(x_stop-x_start), math.ceil(y_stop-y_start), probability]
        new_object_box.append(narrow_box)
    
    return new_object_box

def find_edge_points_one_object(small_img, one_cell_sign_pos):
    # 找到某元件边缘的点
    # 返回元件边缘点的集合
    # small_img比识别框大一圈(min_length_line/2)
    edge_points = {}
    
    def merge_near_point(points):
        if len(points)>=1:
            points = np.append(points, 0)
            index_list = np.where(np.diff(points)>min_length_line)[0]
            index_temp = [x+1 for x in index_list]
            index_temp.insert(0, 0)
            index_temp.insert(len(index_temp), -1)

            return_result = [points[index_temp[index]:index_temp[index+1]] for index in range(len(index_temp)-1)]
            return_result = [[np.min(np.array(x)), np.max(np.array(x))] for x in return_result if (max(x)-min(x))<min(small_img.shape)*1]
        else:
            return_result = []

        return return_result

    min_x, min_y, max_x, max_y = one_cell_sign_pos
    # _, bin_img = cv2.threshold(small_img, 220, 255, cv2.THRESH_BINARY)
    bin_img = small_img
    add_one_position = []
    for direction in ['E', 'S', 'W', 'N']:
        edge_points[direction] = []
        if direction=='E' and np.all(bin_img)!=None:
            select_data = bin_img[round(min_length_line//2):min(-round(min_length_line//2), -1), -1-round(min_length_line//2):]
            width_line_mean = np.mean(select_data, axis=1)
            points = np.array([index for index in range(len(width_line_mean)) if width_line_mean[index]==0])
            points = merge_near_point(points)
            edge_points[direction] = [[[max_x, y_up+min_y], [max_x, y_down+min_y]] for y_up,y_down in points]
            if len(edge_points[direction])==1:
                opposite_direction_position = [[[min_x, y_up+min_y], [min_x, y_down+min_y]] for y_up,y_down in points]
                add_one_position.append(['W', opposite_direction_position ])
        elif direction=='W' and np.all(bin_img)!=None:
            select_data = bin_img[round(min_length_line//2):min(-round(min_length_line//2), -1), :round(min_length_line//2)]
            width_line_mean = np.mean(select_data, axis=1)
            points = np.array([index for index in range(len(width_line_mean)) if width_line_mean[index]==0])
            points = merge_near_point(points)
            edge_points[direction] = [[[min_x, y_up+min_y], [min_x, y_down+min_y]] for y_up,y_down in points]
            if len(edge_points[direction])==1:
                opposite_direction_position = [[[max_x, y_up+min_y], [max_x, y_down+min_y]] for y_up,y_down in points]
                add_one_position.append(['E', opposite_direction_position])

        elif direction=='S' and np.all(bin_img)!=None:
            select_data = bin_img[-1-round(min_length_line//2):, round(min_length_line//2):min(-round(min_length_line//2), -1)]
            height_line_mean = np.mean(select_data, axis=0)
            points = np.array([index for index in range(len(height_line_mean)) if height_line_mean[index]==0])
            points = merge_near_point(points)
            edge_points[direction] = [[[x_left+min_x, max_y], [x_right+min_x, max_y]] for x_left,x_right in points]
            if len(edge_points[direction])==1:
                opposite_direction_position = [[[x_left+min_x, min_y], [x_right+min_x, min_y]] for x_left,x_right in points]
                add_one_position.append(['N', opposite_direction_position])
        elif direction=='N' and np.all(bin_img)!=None:
            select_data = bin_img[:round(min_length_line//2), round(min_length_line//2):min(-round(min_length_line//2), -1)]
            height_line_mean = np.mean(select_data, axis=0)
            points = np.array([index for index in range(len(height_line_mean)) if height_line_mean[index]==0])
            points = merge_near_point(points)
            edge_points[direction] = [[[x_left+min_x, min_y], [x_right+min_x, min_y]] for x_left,x_right in points]
            if len(edge_points[direction])==1:
                opposite_direction_position = [[[x_left+min_x, max_y], [x_right+min_x, max_y]] for x_left,x_right in points]
                add_one_position.append(['S', opposite_direction_position])

    # # 单端点的拐点，将其对立面增加为连接点
    # if len(add_one_position)==1:
    #     direction = add_one_position[0][0]
    #     connect_position = add_one_position[0][1]
    #     edge_points[direction] = connect_position
    
    # small_img = cv2.cvtColor(small_img, cv2.COLOR_GRAY2RGB)
    # for direction in ['E', 'S', 'W', 'N']:
    #     corner_list = edge_points[direction]
    #     for corner in corner_list:
    #         cv2.line(small_img, (int(corner[0][0]-min_x), int(corner[0][1]-min_y)), (int(corner[1][0]-min_x), int(corner[1][1]-min_y)), (255,0,0), 2)
    
    # cv2.imshow('ss', small_img)
    # cv2.waitKey(5000)

    return edge_points

def read_csv_data(file_path):
    # 读取固定格式的表格数据
    wb = load_workbook(file_path)
    sheet_data = wb['Sheet1']
    row = sheet_data.max_row
    column = sheet_data.max_column
    data = []
    for i in range(2,row+1):
        data_temp = [sheet_data.cell(i, x).value for x in range(1, column+1)]
        # data_temp[0] = int(data_temp[0].split('_')[1])
        data_temp[0] = int(data_temp[0])
        data_temp[1] = data_temp[1]+data_temp[3]/2 
        data_temp[2] = data_temp[2]+data_temp[4]/2 
        data_temp.append(1.0)
        data.append(data_temp)
    
    return list(np.array(data, dtype=np.int))

def get_config_parameter(config_path):
    # 读取测试程序配置文件
    with open(config_path, 'r') as json_f:
        parameter_info = json.load(json_f)
    
    return parameter_info


def save_reco_info(object_list, file_name, save_path):
    all_info = {}
    for object in object_list:
        object_num = object.object_num
        all_info[object_num] = {}
        all_info[object_num]['label'] = object.label
        all_info[object_num]['no'] = object.no
        all_info[object_num]['type'] = object.type
        all_info[object_num]['x_center'] = object.x_center
        all_info[object_num]['y_center'] = object.y_center
        all_info[object_num]['width'] = object.width
        all_info[object_num]['height'] = object.height
        all_info[object_num]['probability'] = object.probability
        all_info[object_num]['show'] = object.show
        all_info[object_num]['chinese_name'] = object.chinese_name
        all_info[object_num]['description'] = object.description
        all_info[object_num]['text'] = object.text

        if object.type in ['cell', 'conn']:
            all_info[object_num]['rectangle'] = object.rectangle
            all_info[object_num]['has_tab'] = object.has_tab
            all_info[object_num]['identify_by_tab'] = object.identify_by_tab
            all_info[object_num]['height_width'] = object.height_width
            all_info[object_num]['keywords'] = object.keywords
            all_info[object_num]['flow_direction'] = object.flow_direction
            all_info[object_num]['link_direction'] = object.link_direction
            all_info[object_num]['link_position'] = object.link_position
            all_info[object_num]['object_num'] = object.object_num
            all_info[object_num]['link_probability'] = object.link_probability
            all_info[object_num]['has_confirm_all_link'] = object.has_confirm_all_link
            # all_info[object_num]['possible_direction_num'] = object.possible_direction_num
            all_info[object_num]['confirm_link_object'] = [[x[0].object_num, x[1], x[2], x[3], x[4], x[5]] for x in object.inference_confirm_link_object]
            all_info[object_num]['unsure_link_object'] = [[x[0].object_num, x[1], x[2], x[3], x[4], x[5]] for x in object.inference_unsure_link_object]
            # all_info[object_num]['possible_link_object'] = [[x[0].object_num, x[1], x[2], x[3], x[4], x[5]] for x in object.inference_link_object]
    
    name, suffix = os.path.splitext(os.path.split(file_name)[1])
    save_result_path = os.path.join(save_path, 'connect_info_'+name+'.json')
    with open(save_result_path, 'w', encoding='utf-8') as json_f:
        json.dump(all_info, json_f, ensure_ascii=False, indent=4)


if __name__=="__main__":
    pass
    # # 功能1测试，计算元件的方向
    # label = 26
    # temperate_parent_path = r'../data/sample/direction_template/{}'.format(label)
    # temperate_paths = [os.path.join(temperate_parent_path, x) for x in os.listdir(temperate_parent_path) if os.path.isfile(os.path.join(temperate_parent_path, x))]
    # detect_img_path = r'../data/sample/direction_template/{}.png'.format(label)
    # direction = get_direction(detect_img_path, temperate_paths)
    # print(direction)

    # # 功能2测试，截取labelme标记的图片保存,单个图片保存
    # img_path = r'../data/005.png'
    # save_img_path = r'../data/sample/label'
    # label_json_path = r'../data/005.json'
    # save_label_img(img_path, save_img_path, label_json_path)

    # # 功能3测试， pdf转图片
    # pdf_path = r'E:\SynologyDrive\project\lct_recon\资料\原文件\TFH-高压给水加热器系统-第10章流程图-A-CFC.pdf'
    # img_path = r'E:\SynologyDrive\project\lct_recon\资料\原文件\转图片'
    # zoom_x = 3
    # zoom_y = 3
    # pdf_image(pdf_path, img_path, zoom_x, zoom_y, rotation_angle=0)

    # # 功能4测试，从标记的单个图标区域周围截取一定范围的图片，用于下一步的标记
    # img_path = r'..\data\sample\used_for_label\one_labelme\022.png'
    # labelme_path = r'..\data\sample\used_for_label\one_labelme\022.json'
    # crop_width = 1000
    # crop_height = 1000
    # save_path = r'..\data\sample\used_for_label'
    # crop_img_for_labelme(img_path, labelme_path, crop_width, crop_height, save_path)

    # # 功能5测试，图像细化
    # # for file_name in ['01_M','02','03','05','06','07','08','09','10','11','12','13','14','15','16','17','18']:
    # for file_name in ['05']:
    #     img_path = r'..\data\sample\labelme\one_labelme\0{}.png'.format(file_name)
    #     img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
    #     thin_line(img, show=True)

    # # 功能6测试, 角点检测
    # img_path = r'..\data\sample\labelme\one_labelme\001_M.png'
    # img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
    # find_corner(img, show=True)

    # # 功能7测试, 元件连接点检测
    # # for file_name in ['01_M','02','03','05','06','07','08','09','10','11','12','13','14','15','16','17','18']:
    # for file_name in ['05']:
    #     all_objects = read_labelme_json(r'..\data\sample\labelme\one_labelme\005.json')
    #     big_img = cv2.imread(r'..\data\sample\labelme\one_labelme\0{}.png'.format(file_name), cv2.IMREAD_GRAYSCALE)
    #     config_path = r'data/config_new.json'
    #     temperate_parent_path = r'..\data/sample/direction_template'
    #     big_img_ori = copy.deepcopy(big_img)
    #     objects_list = build_objects_info(all_objects, big_img, config_path)
    #     no_cell_sign_img, cell_sign_pos_list = big_img_remove_object_space(big_img, objects_list)
    #     no_cell_sign_img = thin_line(no_cell_sign_img)
    #     corner_list = find_corner(no_cell_sign_img)
    #     find_edge_points(big_img_ori, corner_list, cell_sign_pos_list, show=True)

    # # 功能8测试, 图像删除元件
    # all_objects = read_labelme_json(r'../data/005.json')
    # big_img = cv2.imread(r'../data/005.png', cv2.IMREAD_GRAYSCALE)
    # config_path = r'data/config_new.json'
    # temperate_parent_path = r'../data/sample/direction_template'
    # objects_list = build_objects_info(all_objects, big_img, config_path)
    # with open('objects_list.pkl', 'wb') as f:
    #     # objects_list = pickle.load(f)
    #     pickle.dump(objects_list, f)
    # new_big_img = big_img_remove_object_space(big_img, objects_list, show=True)

    # # 功能9测试,拐点检测
    # # for file_name in ['01_M','02','03','05','06','07','08','09','10','11','12','13','14','15','16','17','18']:
    # for file_name in ['06']:
    #     all_objects = read_labelme_json(r'..\data\sample\labelme\one_labelme\011.json')
    #     big_img = cv2.imread(r'..\data\sample\labelme\one_labelme\0{}.png'.format(file_name), cv2.IMREAD_GRAYSCALE)
    #     config_path = r'data\config_new.json'
    #     temperate_parent_path = r'..\data\sample\direction_template'
    #     objects_list = build_objects_info(all_objects, big_img, config_path)
    #     big_img_ori = copy.deepcopy(big_img)
    #     no_cell_sign_img, _ = big_img_remove_object_space(big_img, objects_list)
    #     thin_line_img = thin_line(no_cell_sign_img)
    #     find_conn(thin_line_img, big_img_ori, file_name, show=True)